import glob
import os
import re

import openpyxl
import pdfplumber
from openpyxl.styles import Font, PatternFill

from nclt_utils import (
    extract_case_references as shared_extract_case_references,
    extract_party_from_cells as shared_extract_party_from_cells,
    find_latest_file,
    normalize_company_key,
    repair_doubled_ocr,
)


PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(PROJECT_DIR, "downloads")
AUDIT_SOURCE_PREFIX = "[AUDIT RECOVERED SOURCE:"
AUDIT_ROW_MARKER = "AUDIT RECOVERED"
AUDIT_REVIEW_SHEET = "Audit Needs Review"

LINE_TABLE_SETTINGS = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
}
TEXT_TABLE_SETTINGS = {
    "vertical_strategy": "text",
    "horizontal_strategy": "text",
    "intersection_x_tolerance": 20,
    "snap_tolerance": 5,
}

VS_PATTERN = re.compile(
    r"(?<![A-Z0-9])(?:(?:V\s*/\s*S|V\s*S|VERSUS)(?:\.)?|V\.)(?![A-Z0-9])",
    re.IGNORECASE,
)
CASE_REFERENCE_PATTERNS = (
    re.compile(
        r"""
        (?<![A-Z0-9])
        C\s*\.?\s*P\s*\.?
        \s*(?:NO\.?\s*)?
        \(\s*I\s*B\s*\)
        \s*(?:NO\.?\s*)?
        (?:[/:\-]\s*)?
        \d+[A-Z]?
        (?:
            \s*/\s*[A-Z0-9.\-]+
            |
            \s*\(\s*[A-Z0-9.\-]+\s*\)
        ){0,5}
        \s*(?:/\s*)?(?:OF\s*)?(?:19|20)\d{2}\b
        """,
        re.IGNORECASE | re.VERBOSE,
    ),
    re.compile(
        r"""
        (?<![A-Z0-9(/])
        (?:IB|TCP|TP)
        \s*\)?\s*(?:NO\.?\s*)?
        [/:\-]\s*
        \d+[A-Z]?
        (?:
            \s*/\s*[A-Z0-9.\-]+
            |
            \s*\(\s*[A-Z0-9.\-]+\s*\)
        ){0,5}
        \s*(?:/\s*)?(?:OF\s*)?(?:19|20)\d{2}\b
        """,
        re.IGNORECASE | re.VERBOSE,
    ),
    re.compile(
        r"""
        (?<![A-Z0-9])
        C\s*\.?\s*P\s*\.?
        \s*(?:NO\.?\s*)?
        [/:\-]\s*\d+[A-Z]?
        \s*/\s*IB
        (?:\s*/\s*[A-Z0-9.\-]+){0,4}
        \s*/\s*(?:19|20)\d{2}\b
        """,
        re.IGNORECASE | re.VERBOSE,
    ),
    re.compile(
        r"""
        (?<![A-Z0-9])
        \(\s*IB\s*\)
        \s*[/:\-]\s*\d+[A-Z]?
        (?:
            \s*/\s*[A-Z0-9.\-]+
            |
            \s*\(\s*[A-Z0-9.\-]+\s*\)
        ){0,4}
        \s*(?:/\s*)?(?:19|20)\d{2}\b
        """,
        re.IGNORECASE | re.VERBOSE,
    ),
)
BLOCK_START_PATTERN = re.compile(
    r"^(?:\d+[A-Z]?\.?|C\s*\.?\s*P|IB|I\s*\.?\s*A|M\s*\.?\s*A|"
    r"C\s*\.?\s*A|TCP|TP|SUPPLEMENTARY|ITEM)\b",
    re.IGNORECASE,
)

BLOCKED_PARTY_WORDS = {
    "ADMITTED",
    "ADV",
    "ADVOCATE",
    "ALLOWED",
    "CONSIDERATION",
    "COUNSEL",
    "IRP",
    "LIQUIDATOR",
    "MATTER",
    "PENDING",
    "RESOLUTION PROFESSIONAL",
    "RP",
}
ENTITY_WORDS = {
    "BANK",
    "COMPANY",
    "CORPORATION",
    "DEVELOPERS",
    "ENTERPRISES",
    "FINANCE",
    "INDIA",
    "INDUSTRIES",
    "INFRA",
    "LIMITED",
    "LLP",
    "LTD",
    "PRIVATE",
    "PVT",
    "SERVICES",
}


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", repair_doubled_ocr(value)).strip()


def normalize_text(value):
    return re.sub(r"[^A-Z0-9]+", " ", clean_text(value).upper()).strip()


def normalize_party_side(value):
    normalized = normalize_text(value)
    normalized = re.sub(r"\bPVT\b", "PRIVATE", normalized)
    normalized = re.sub(r"\bLTD\b", "LIMITED", normalized)
    normalized = re.sub(r"\bM S\b", "", normalized)
    return clean_text(normalized)


def get_latest_master_file():
    return find_latest_file("Consolidated_*Master_*.xlsx")


def get_download_folder_for_master(master_file):
    timestamp_match = re.search(
        r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})",
        os.path.basename(master_file),
    )
    if timestamp_match:
        matching_folder = os.path.join(DOWNLOAD_DIR, timestamp_match.group(1))
        return matching_folder if os.path.isdir(matching_folder) else None

    if not os.path.isdir(DOWNLOAD_DIR):
        return None
    folders = [
        os.path.join(DOWNLOAD_DIR, name)
        for name in os.listdir(DOWNLOAD_DIR)
        if os.path.isdir(os.path.join(DOWNLOAD_DIR, name))
    ]
    return max(folders, key=os.path.getctime) if folders else None


def is_section_7(text):
    """Detect Section 7 applications without treating every Rule 7 as IBC Section 7."""
    if not text:
        return False
    value = clean_text(text).upper()
    explicit_patterns = [
        r"\b(?:SEC(?:TION)?|U/S|U\\S|S\.)\s*[-.:]?\s*7\b",
        r"\bIBC\s*(?:UNDER\s*)?(?:SEC(?:TION)?\s*)?7\b",
    ]
    if any(re.search(pattern, value) for pattern in explicit_patterns):
        return True
    if re.search(r"\bRULE\s*[-.:]?\s*7\b", value):
        return False
    implicit_patterns = [
        r"\b7\s*(?:OF\s+)?(?:THE\s+)?(?:IBC|INSOLVENCY|CODE)\b",
        r"\b7\s*IBC\b",
    ]
    return any(re.search(pattern, value) for pattern in implicit_patterns)


def extract_case_details(text):
    """Extract complete NCLT case references that include a case number and year."""
    return [
        {"key": reference["key"], "display": reference["display"]}
        for reference in shared_extract_case_references(text)
    ]


def split_party_text(text):
    source = clean_text(text)
    match = VS_PATTERN.search(source)
    if not match:
        return "", ""
    return clean_text(source[: match.start()]), clean_text(source[match.end() :])


def plausible_party_side(value):
    normalized = normalize_party_side(value)
    words = normalized.split()
    if len(words) < 2:
        return False
    if any(
        re.search(rf"\b{re.escape(blocked)}\b", normalized)
        for blocked in BLOCKED_PARTY_WORDS
    ):
        return False
    return sum(character.isalpha() for character in normalized) >= 6


def party_score(text):
    creditor, debtor = split_party_text(text)
    normalized = f"{normalize_party_side(creditor)} {normalize_party_side(debtor)}"
    entity_matches = sum(
        1 for word in ENTITY_WORDS if re.search(rf"\b{re.escape(word)}\b", normalized)
    )
    return len(normalized) + (entity_matches * 20)


def has_entity_word(value):
    normalized = normalize_party_side(value)
    return any(
        re.search(rf"\b{re.escape(word)}\b", normalized)
        for word in ENTITY_WORDS
    )


def text_table_party_is_safe(text):
    creditor, debtor = split_party_text(text)
    return has_entity_word(creditor) and has_entity_word(debtor)


def extract_party_text(cells):
    """Choose a complete party cell; split table fragments are left for review."""
    parsed = shared_extract_party_from_cells(cells)
    if not parsed or not parsed["confident"]:
        return ""
    return f"{parsed['creditor']} Vs {parsed['debtor']}"


def party_key(text):
    creditor, debtor = split_party_text(text)
    if not creditor or not debtor:
        return ""
    return f"{normalize_party_side(creditor)} VS {normalize_party_side(debtor)}"


def debtor_key(text):
    _, debtor = split_party_text(text)
    return normalize_company_key(debtor) if debtor else ""


def extract_case_details_from_cells(cells):
    details = []
    seen = set()
    for cell in cells:
        for detail in extract_case_details(cell):
            if detail["key"] in seen:
                continue
            details.append(detail)
            seen.add(detail["key"])
    return details


def page_blocks(page):
    raw_text = page.extract_text(layout=True)
    if not raw_text:
        return []

    blocks = []
    current_lines = []
    for raw_line in raw_text.splitlines():
        if not raw_line.strip():
            continue

        clean_line = raw_line.lstrip()
        indentation = len(raw_line) - len(clean_line)
        if indentation < 20 and BLOCK_START_PATTERN.match(clean_line) and current_lines:
            blocks.append(current_lines)
            current_lines = []
        current_lines.append(clean_line)

    if current_lines:
        blocks.append(current_lines)
    return blocks


def explicit_non_section_7_match(text):
    return re.search(
        r"\b(?:(?:IBC\s+)?UNDER\s+)?(?:SEC(?:TION)?\.?|U/S)\s*(?:9|10)\b|"
        r"\b(?:9|10)\s+(?:OF\s+)?IBC\b",
        text,
        re.I,
    )


def raw_block_starts_with_non_ibc_case(text):
    section_7 = re.search(
        r"\b(?:IBC\s+UNDER\s+)?SEC(?:TION)?\.?\s*7\b|\b7\s+OF\s+IBC\b",
        text,
        re.I,
    )
    company_law = re.search(
        r"\bSEC(?:TION)?\.?\s*(?:230|232|241|242|244)\b",
        text,
        re.I,
    )
    non_section_7 = explicit_non_section_7_match(text)
    non_ibc = min(
        (match for match in (company_law, non_section_7) if match),
        key=lambda match: match.start(),
        default=None,
    )
    return bool(section_7 and non_ibc and non_ibc.start() < section_7.start())


def make_item(source_file, page_number, cells, extraction_mode):
    case_details = extract_case_details_from_cells(cells)
    parties = extract_party_text(cells)
    return {
        "source_file": source_file,
        "page": page_number,
        "case_id": case_details[0]["display"] if case_details else "",
        "primary_reference": case_details[0]["key"] if case_details else "",
        "references": {item["key"] for item in case_details},
        "party_text": parties,
        "party_key": party_key(parties),
        "debtor_key": debtor_key(parties),
        "raw_text": " | ".join(clean_text(cell) for cell in cells if clean_text(cell)),
        "extraction_mode": extraction_mode,
    }


def collect_audit_candidates(pdf_files):
    candidate_by_case = {}
    review_items = []
    errors = []
    seen_review = set()

    def add_review(item):
        raw_text = item["raw_text"]
        looks_like_split_case = bool(
            VS_PATTERN.search(raw_text)
            and re.search(r"(?:19|20)\s*\d{2}", raw_text)
        )
        if not item["references"] and not item["party_key"] and not looks_like_split_case:
            return
        review_key = (
            item["source_file"],
            item["page"],
            normalize_text(item["raw_text"]),
        )
        if review_key[2] and review_key not in seen_review:
            review_items.append(item)
            seen_review.add(review_key)

    for pdf_path in pdf_files:
        source_file = os.path.basename(pdf_path)
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_number, page in enumerate(pdf.pages, start=1):
                    page_candidate_references = set()
                    page_non_section_7_references = set()

                    for extraction_mode, settings in (
                        ("line table", LINE_TABLE_SETTINGS),
                        ("text table", TEXT_TABLE_SETTINGS),
                    ):
                        try:
                            tables = page.extract_tables(settings) or []
                        except Exception as exc:
                            errors.append(
                                f"{source_file}, page {page_number}, {extraction_mode}: "
                                f"{type(exc).__name__}: {exc}"
                            )
                            continue

                        for table in tables:
                            for row in table or []:
                                cells = [clean_text(value) for value in (row or [])]
                                row_text = " ".join(cells)
                                row_references = {
                                    detail["key"]
                                    for detail in extract_case_details_from_cells(cells)
                                }
                                if explicit_non_section_7_match(row_text):
                                    page_non_section_7_references.update(row_references)
                                    continue
                                if not is_section_7(row_text):
                                    continue

                                item = make_item(
                                    source_file,
                                    page_number,
                                    cells,
                                    extraction_mode,
                                )
                                automatic_recovery_is_safe = (
                                    extraction_mode == "line table"
                                    or text_table_party_is_safe(item["party_text"])
                                )
                                if (
                                    item["references"]
                                    and item["party_key"]
                                    and automatic_recovery_is_safe
                                ):
                                    candidate_key = (
                                        source_file,
                                        item["primary_reference"],
                                    )
                                    existing = candidate_by_case.get(candidate_key)
                                    if not existing or party_score(item["party_text"]) > party_score(
                                        existing["party_text"]
                                    ):
                                        candidate_by_case[candidate_key] = item
                                    page_candidate_references.update(item["references"])
                                elif not (item["references"] & page_candidate_references):
                                    add_review(item)

                    # Raw layout is retained as a safety net for human review only.
                    for lines in page_blocks(page):
                        block_text = clean_text(" ".join(lines))
                        if not is_section_7(block_text):
                            continue
                        if raw_block_starts_with_non_ibc_case(block_text):
                            continue
                        item = make_item(source_file, page_number, lines, "raw layout")
                        if item["references"] & page_non_section_7_references:
                            continue
                        if item["references"] & page_candidate_references:
                            continue
                        add_review(item)
        except Exception as exc:
            errors.append(f"{source_file}: {type(exc).__name__}: {exc}")

    return list(candidate_by_case.values()), review_items, errors


def load_master_index_details(master_file):
    workbook = openpyxl.load_workbook(master_file, read_only=True, data_only=True)
    worksheet = workbook.active
    references = set()
    parties = set()
    debtors = set()

    for row in worksheet.iter_rows(values_only=True):
        cells = [str(value) for value in row if clean_text(value)]
        if not cells:
            continue
        for detail in extract_case_details_from_cells(cells):
            references.add(detail["key"])
        normalized_party = party_key(extract_party_text(cells))
        if normalized_party:
            parties.add(normalized_party)
            normalized_debtor = debtor_key(normalized_party)
            if normalized_debtor:
                debtors.add(normalized_debtor)

    workbook.close()
    return references, parties, debtors


def load_master_index(master_file):
    references, parties, _ = load_master_index_details(master_file)
    return references, parties


def party_already_present(candidate_party, master_parties):
    if not candidate_party:
        return False
    if candidate_party in master_parties:
        return True
    return any(
        len(existing) >= 20
        and (candidate_party in existing or existing in candidate_party)
        for existing in master_parties
    )


def debtor_already_present(candidate_debtor, master_debtors):
    if not candidate_debtor:
        return False
    if candidate_debtor in master_debtors:
        return True
    return any(
        len(existing) >= 8
        and (candidate_debtor in existing or existing in candidate_debtor)
        for existing in master_debtors
    )


def reference_core(reference):
    core = re.sub(r"^(?:R?CP)", "", str(reference or "").upper())
    return re.sub(r"^IBC", "IB", core)


def references_already_present(candidate_references, master_references):
    if candidate_references & master_references:
        return True
    master_cores = {reference_core(item) for item in master_references}
    return any(reference_core(item) in master_cores for item in candidate_references)


def find_missing_candidates(
    candidates,
    master_references,
    master_parties,
    master_debtors=None,
):
    master_debtors = master_debtors or set()
    missing = []
    for candidate in candidates:
        reference_match = references_already_present(
            candidate["references"],
            master_references,
        )
        party_match = party_already_present(candidate["party_key"], master_parties)
        debtor_match = debtor_already_present(candidate["debtor_key"], master_debtors)
        if not reference_match and not party_match and not debtor_match:
            missing.append(candidate)
    return missing


def filter_review_items(review_items, master_references, master_parties, master_debtors):
    unresolved = []
    seen = set()
    for item in review_items:
        if references_already_present(item["references"], master_references):
            continue
        if party_already_present(item["party_key"], master_parties):
            continue
        if debtor_already_present(item["debtor_key"], master_debtors):
            continue
        if not item["references"] and not item["debtor_key"]:
            continue
        if item["extraction_mode"] != "line table" and not item["references"]:
            continue

        key = (
            tuple(sorted(item["references"])),
            item["debtor_key"],
            item["source_file"],
        )
        if key in seen:
            continue
        unresolved.append(item)
        seen.add(key)
    return unresolved


def append_recoveries(master_file, missing_cases, review_items):
    workbook = openpyxl.load_workbook(master_file)
    worksheet = workbook.active
    max_columns = max(worksheet.max_column, 7)
    banner_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    recovered_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    grouped = {}
    for candidate in missing_cases:
        grouped.setdefault(candidate["source_file"], []).append(candidate)

    for source_file, source_cases in grouped.items():
        banner = [f"{AUDIT_SOURCE_PREFIX} {source_file}]"] + [""] * (max_columns - 1)
        worksheet.append(banner)
        for cell in worksheet[worksheet.max_row]:
            cell.fill = banner_fill
            cell.font = Font(bold=True)

        for candidate in source_cases:
            recovered_row = [
                AUDIT_ROW_MARKER,
                candidate["case_id"],
                candidate["source_file"],
                f"Page {candidate['page']}",
                "Section 7 IBC",
                candidate["party_text"],
                candidate["raw_text"],
            ]
            recovered_row.extend([""] * (max_columns - len(recovered_row)))
            worksheet.append(recovered_row)
            worksheet.cell(worksheet.max_row, 1).fill = recovered_fill
        worksheet.append([""] * max_columns)

    if AUDIT_REVIEW_SHEET in workbook.sheetnames:
        workbook.remove(workbook[AUDIT_REVIEW_SHEET])

    newly_logged_reviews = 0
    if review_items:
        review_sheet = workbook.create_sheet(AUDIT_REVIEW_SHEET)
        review_sheet.append(
            ["Source PDF", "Page", "Detected Case ID", "Party Text", "Raw Text", "Reason"]
        )
        for cell in review_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = banner_fill

        seen_fingerprints = set()
        for candidate in review_items:
            fingerprint = normalize_text(candidate["raw_text"])
            if not fingerprint or fingerprint in seen_fingerprints:
                continue
            reason = []
            if not candidate["references"]:
                reason.append("No complete case number and year")
            if not candidate["party_key"]:
                reason.append("No complete creditor Vs debtor cell")
            reason.append(f"Source: {candidate['extraction_mode']}")
            review_sheet.append(
                [
                    candidate["source_file"],
                    candidate["page"],
                    candidate["case_id"],
                    candidate["party_text"],
                    candidate["raw_text"],
                    "; ".join(reason),
                ]
            )
            seen_fingerprints.add(fingerprint)
            newly_logged_reviews += 1

    temporary_path = f"{master_file}.audit_tmp.xlsx"
    output_path = master_file
    try:
        try:
            workbook.save(temporary_path)
        finally:
            workbook.close()

        try:
            os.replace(temporary_path, master_file)
        except PermissionError as exc:
            # Excel and preview handlers can lock the original workbook. Keep
            # it untouched and hand the next pipeline stage a fresh copy.
            base, extension = os.path.splitext(master_file)
            output_path = f"{base}_audit_recovered{extension}"
            suffix = 2
            while os.path.exists(output_path):
                output_path = f"{base}_audit_recovered_{suffix}{extension}"
                suffix += 1

            try:
                os.replace(temporary_path, output_path)
            except PermissionError as fallback_exc:
                raise RuntimeError(
                    "The master workbook is locked by another application. "
                    "Close it in Excel or the Windows preview pane and rerun audit.py. "
                    f"The original workbook was left unchanged: {master_file}"
                ) from fallback_exc

            print(
                "[WARN] Master workbook is locked; the original was left unchanged. "
                f"Audit output saved to: {output_path}"
            )
    finally:
        if os.path.exists(temporary_path):
            os.remove(temporary_path)

    return len(missing_cases), newly_logged_reviews, output_path


def run_audit():
    print("\n==================================================")
    print("   AUDIT: RECOVERING MISSED IBC SECTION 7 CASES")
    print("==================================================\n")

    master_file = get_latest_master_file()
    if not master_file:
        raise RuntimeError("No Consolidated Master workbook found. Run scraper.py first.")

    download_folder = get_download_folder_for_master(master_file)
    if not download_folder:
        raise RuntimeError("No matching download folder found for the Master workbook.")

    pdf_files = sorted(glob.glob(os.path.join(download_folder, "*.pdf")))
    if not pdf_files:
        raise RuntimeError(f"No PDFs found in {download_folder}.")

    print(f"[*] Master workbook : {master_file}")
    print(f"[*] PDF run folder  : {os.path.basename(download_folder)}")
    print(f"[*] PDFs to audit   : {len(pdf_files)}")

    candidates, review_items, errors = collect_audit_candidates(pdf_files)
    if errors:
        raise RuntimeError("Audit could not read every PDF: " + " | ".join(errors))

    master_references, master_parties, master_debtors = load_master_index_details(
        master_file
    )
    missing_cases = find_missing_candidates(
        candidates,
        master_references,
        master_parties,
        master_debtors,
    )
    unresolved_reviews = filter_review_items(
        review_items,
        master_references,
        master_parties,
        master_debtors,
    )
    appended_count, review_count, audit_output = append_recoveries(
        master_file,
        missing_cases,
        unresolved_reviews,
    )

    print("\n==================================================")
    print("AUDIT RECOVERY COMPLETE")
    print("==================================================")
    print(f"Validated Section 7 candidates : {len(candidates)}")
    print(f"Already present in Master      : {len(candidates) - len(missing_cases)}")
    print(f"Recovered into Master          : {appended_count}")
    print(f"New ambiguous items logged     : {review_count}")
    print(f"Audit workbook                 : {audit_output}")

    if appended_count:
        print("\n[OK] Missing cases were appended to the Master workbook before CIN enrichment.")
    else:
        print("\n[OK] No validated cases were missing from the Master workbook.")
    if review_count:
        print(f"[REVIEW] See the '{AUDIT_REVIEW_SHEET}' sheet for ambiguous text blocks.")

    return appended_count


if __name__ == "__main__":
    run_audit()
