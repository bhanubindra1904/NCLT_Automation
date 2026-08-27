import glob
import json
import os
import re
import time
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nclt_utils import configured_output_directory, latest_output_directory

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    webdriver = None
    Options = None
    Service = None
    By = None
    Keys = None
    EC = None
    WebDriverWait = None
    ChromeDriverManager = None


PRIVATECIRCLE_MCA_URL = "https://privatecircle.co/company/mca_new/mca-listing/"
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIGURED_OUTPUT_DIR = os.getenv("NCLT_OUTPUT_DIR", "").strip()
OUTPUT_DIR = (
    configured_output_directory()
    if CONFIGURED_OUTPUT_DIR
    else latest_output_directory("CIN_Directory_*.xlsx") or os.getcwd()
)
OUTPUT_FOLDER_TIMESTAMP = os.path.basename(os.path.normpath(OUTPUT_DIR))
RUN_TIMESTAMP = (
    os.getenv("NCLT_RUN_TIMESTAMP", "").strip()
    or (
        OUTPUT_FOLDER_TIMESTAMP
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}", OUTPUT_FOLDER_TIMESTAMP)
        else datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    )
)
DEBUG_DIR = os.path.join(PROJECT_DIR, "privatecircle_debug")
SAVE_DEBUG = os.getenv("PRIVATECIRCLE_SAVE_DEBUG", "0").strip().lower() in {"1", "true", "yes"}

MAX_COMPANIES = int(os.getenv("PRIVATECIRCLE_MAX_COMPANIES", "0") or "0")
SEARCH_WAIT_SECONDS = float(os.getenv("PRIVATECIRCLE_SEARCH_WAIT_SECONDS", "4") or "4")
SEARCH_RESULT_TIMEOUT_SECONDS = float(os.getenv("PRIVATECIRCLE_SEARCH_RESULT_TIMEOUT_SECONDS", "25") or "25")
SEARCH_ATTEMPTS = int(os.getenv("PRIVATECIRCLE_SEARCH_ATTEMPTS", "2") or "2")
PROFILE_LOAD_TIMEOUT_SECONDS = float(os.getenv("PRIVATECIRCLE_PROFILE_LOAD_TIMEOUT_SECONDS", "45") or "45")
FINANCIAL_LOAD_TIMEOUT_SECONDS = float(os.getenv("PRIVATECIRCLE_FINANCIAL_LOAD_TIMEOUT_SECONDS", "45") or "45")
FINANCIAL_CAPTURE_ATTEMPTS = int(os.getenv("PRIVATECIRCLE_FINANCIAL_CAPTURE_ATTEMPTS", "2") or "2")
FINANCIAL_UNCHANGED_GRACE_SECONDS = float(os.getenv("PRIVATECIRCLE_FINANCIAL_UNCHANGED_GRACE_SECONDS", "5") or "5")
COMPANY_ATTEMPTS = int(os.getenv("PRIVATECIRCLE_COMPANY_ATTEMPTS", "2") or "2")
LOGIN_TIMEOUT_SECONDS = float(os.getenv("PRIVATECIRCLE_LOGIN_TIMEOUT_SECONDS", "900") or "900")
LOGIN_POLL_SECONDS = float(os.getenv("PRIVATECIRCLE_LOGIN_POLL_SECONDS", "2") or "2")
CHECKPOINT_EVERY = int(os.getenv("PRIVATECIRCLE_CHECKPOINT_EVERY", "1") or "1")


OUTPUT_COLUMNS = [
    "Court",
    "Case ID",
    "Financial Creditor",
    "Corporate Debtor",
    "CIN / LLPIN",
    "Company Status",
    "CIRP Status",
    "Active Under CIRP",
    "Financial Basis",
    "Financial Year",
    "Financial Units",
    "Revenue",
    "EBITDA",
    "Long Term Debt",
    "Short Term Debt",
    "Debt (LT + ST)",
    "Inventory",
    "Fixed Assets",
    "Asset (Inventory + Fixed)",
    "PrivateCircle URL",
    "Extraction Status",
    "Notes",
]


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


WINDOWS_LOCAL_PATH_RE = re.compile(
    r'(?i)\b[A-Z]:\\(?:[^\\/:*?"<>|\r\n]+\\)*[^\\/:*?"<>|\r\n]*'
)
UNIX_HOME_PATH_RE = re.compile(r"(?i)(?<!\w)/(?:Users|home)/[^\r\n]*")


def redact_local_paths(value):
    text = clean_text(value)
    text = WINDOWS_LOCAL_PATH_RE.sub("[local path removed]", text)
    text = UNIX_HOME_PATH_RE.sub("[local path removed]", text)
    return clean_text(text)


def norm(value):
    return re.sub(r"[^A-Z0-9]+", " ", clean_text(value).upper()).strip()


def get_latest_cin_file():
    files = glob.glob(os.path.join(OUTPUT_DIR, "CIN_Directory_*.xlsx"))
    if not files and not CONFIGURED_OUTPUT_DIR:
        files = glob.glob(os.path.join(PROJECT_DIR, "CIN_Directory_*.xlsx"))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def pick_column(df, candidates):
    exact = {clean_text(col).lower(): col for col in df.columns}
    for candidate in candidates:
        if candidate.lower() in exact:
            return exact[candidate.lower()]

    for col in df.columns:
        name = clean_text(col).lower()
        if any(candidate.lower() in name for candidate in candidates):
            return col
    return None


def load_cin_rows(file_path):
    df = pd.read_excel(file_path)

    company_col = pick_column(df, ["Corporate Debtor", "Company", "Debtor"])
    cin_col = pick_column(df, ["CIN / LLPIN", "CIN", "LLPIN"])
    court_col = pick_column(df, ["Court"])
    case_col = pick_column(df, ["Case ID", "Case"])
    creditor_col = pick_column(df, ["Financial Creditor", "Creditor"])

    if not company_col:
        raise ValueError("Could not find a Corporate Debtor column in the CIN workbook.")
    if not cin_col:
        raise ValueError("Could not find a CIN / LLPIN column in the CIN workbook.")

    rows = []
    seen = set()
    for _, row in df.iterrows():
        company = clean_text(row.get(company_col, ""))
        cin = clean_text(row.get(cin_col, ""))
        if not company or not cin or cin.upper() in {
            "N/A",
            "NA",
            "NOT FOUND",
            "ERROR",
            "PARSE REVIEW",
        }:
            continue

        key = (company.upper(), cin.upper())
        if key in seen:
            continue
        seen.add(key)

        rows.append(
            {
                "Court": clean_text(row.get(court_col, "")) if court_col else "",
                "Case ID": clean_text(row.get(case_col, "")) if case_col else "",
                "Financial Creditor": clean_text(row.get(creditor_col, "")) if creditor_col else "",
                "Corporate Debtor": company,
                "CIN / LLPIN": cin,
            }
        )

    if MAX_COMPANIES > 0:
        rows = rows[:MAX_COMPANIES]
    return rows


def setup_driver():
    if webdriver is None:
        raise RuntimeError("Install Selenium dependencies first: pip install selenium webdriver-manager")

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("detach", True)
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)


def wait_for_body(driver, timeout=30):
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.TAG_NAME, "body")))


def safe_current_url(driver):
    try:
        return clean_text(driver.current_url)
    except Exception:
        return ""


def page_needs_login(driver):
    try:
        page_text = driver.find_element(By.TAG_NAME, "body").text
    except Exception:
        return False
    return "Username or Email Address" in page_text or "Password" in page_text


def wait_for_manual_login(driver):
    print("[LOGIN] Opening PrivateCircle MCA listing page...")
    driver.get(PRIVATECIRCLE_MCA_URL)
    wait_for_body(driver)

    if page_needs_login(driver):
        print("\n[LOGIN REQUIRED]")
        print("1. Log in to PrivateCircle in the Chrome window that just opened.")
        print("2. Complete OTP/captcha yourself if it appears.")
        print("3. No need to press ENTER. This script will start automatically once the Company Master List is ready.")

    end_time = time.time() + LOGIN_TIMEOUT_SECONDS
    attempted_auto_navigation = False
    manual_navigation_message_shown = False

    while time.time() < end_time:
        try:
            wait_for_body(driver, timeout=10)
        except Exception:
            time.sleep(LOGIN_POLL_SECONDS)
            continue

        if page_needs_login(driver):
            time.sleep(LOGIN_POLL_SECONDS)
            continue

        if is_mca_listing_url(driver.current_url) and not looks_like_dashboard(driver):
            print("[LOGIN] Company Master List loaded. Starting enrichment...")
            return

        if not attempted_auto_navigation:
            attempted_auto_navigation = True
            try:
                navigate_to_mca_listing(driver)
                print("[LOGIN] Company Master List loaded. Starting enrichment...")
                return
            except RuntimeError:
                pass

        try:
            if looks_like_dashboard(driver) and click_company_master_list_link(driver):
                time.sleep(3)
                continue
        except Exception:
            pass

        if not manual_navigation_message_shown:
            manual_navigation_message_shown = True
            print("\n[WAITING]")
            print("PrivateCircle is logged in, but Company Master List is not visible yet.")
            print("Open Companies -> Company Master List in Chrome. The script will continue automatically.")

        time.sleep(LOGIN_POLL_SECONDS)

    raise RuntimeError(
        "Timed out waiting for PrivateCircle Company Master List. "
        "Increase PRIVATECIRCLE_LOGIN_TIMEOUT_SECONDS if login/OTP takes longer."
    )


def is_mca_listing_url(url):
    return "/company/mca_new/mca-listing" in urlparse(url).path


def is_company_detail_url(url):
    path = urlparse(url).path.lower()
    if "/company/" not in path:
        return False
    return not any(
        blocked in path
        for blocked in [
            "/company/mca_new/mca-listing",
            "/company/company_new/comp-listing",
            "dashboard",
            "login",
            "logout",
        ]
    )


def is_full_company_profile_url(url):
    path = urlparse(url).path.lower()
    return any(
        marker in path
        for marker in [
            "/company/profile_updated/",
            "/company/profile/",
        ]
    )


def is_master_company_profile_url(url):
    return "/company/master_profile_updated/" in urlparse(url).path.lower()


def looks_like_dashboard(driver):
    try:
        title = clean_text(driver.title).lower()
        url = driver.current_url.lower()
        body = driver.find_element(By.TAG_NAME, "body").text
    except Exception:
        return False

    body_upper = body.upper()
    return (
        "dashboard" in url
        or title.startswith("dashboard")
        or ("LATEST FOR YOU" in body_upper and "COMPANY MASTER LIST" in body_upper)
    )


def click_company_master_list_link(driver):
    links = driver.find_elements(By.XPATH, "//a[contains(@href, '/company/mca_new/mca-listing')]")
    for link in links:
        try:
            href = link.get_attribute("href")
            if not href:
                continue
            if link.is_displayed() and link.is_enabled():
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", link)
                time.sleep(0.2)
                link.click()
                return True
        except Exception:
            continue

    for link in links:
        href = link.get_attribute("href")
        if href:
            driver.execute_script("window.location.href = arguments[0];", href)
            return True

    return False


def navigate_to_mca_listing(driver):
    driver.get(PRIVATECIRCLE_MCA_URL)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    time.sleep(2)

    if is_mca_listing_url(driver.current_url) and not looks_like_dashboard(driver):
        return

    clicked = click_company_master_list_link(driver)

    if clicked:
        time.sleep(4)
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

    if not is_mca_listing_url(driver.current_url) or looks_like_dashboard(driver):
        raise RuntimeError(
            "Could not reach PrivateCircle Company Master List. "
            "After login, open Company Master List once and rerun."
        )


def page_payload(driver):
    return driver.execute_script(
        """
        const clean = (value) => (value || '').replace(/\\s+/g, ' ').trim();
        const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
        const tables = Array.from(document.querySelectorAll('table')).map((table) => {
          return Array.from(table.querySelectorAll('tr')).map((tr) => {
            return Array.from(tr.querySelectorAll('th,td')).map((cell) => clean(cell.innerText));
          }).filter((row) => row.some(Boolean));
        });

        const gridRows = Array.from(document.querySelectorAll('[role="row"], .ag-row, .data-row')).map((row) => {
          const cells = Array.from(row.querySelectorAll('[role="gridcell"], .ag-cell, td, th')).map((cell) => clean(cell.innerText));
          if (cells.some(Boolean)) return cells;
          const text = clean(row.innerText);
          return text ? [text] : [];
        }).filter((row) => row.some(Boolean));

        if (gridRows.length) tables.push(gridRows);

        const inputs = Array.from(document.querySelectorAll('input, textarea')).map((el) => ({
          tag: el.tagName,
          type: el.getAttribute('type') || '',
          name: el.getAttribute('name') || '',
          id: el.id || '',
          placeholder: el.getAttribute('placeholder') || '',
          visible: !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)
        }));

        const links = Array.from(document.querySelectorAll('a[href]')).slice(0, 500).map((a) => ({
          text: clean(a.innerText),
          href: a.href
        }));

        const controlSelector = [
          'a',
          'button',
          'label',
          '[role="tab"]',
          '[role="button"]',
          '[role="menuitem"]',
          '[role="option"]',
          '[onclick]',
          '[data-toggle="tab"]',
          '[data-bs-toggle="tab"]',
          '.mat-tab-label',
          '.mat-option',
          '.mat-button-toggle-button',
          'mat-button-toggle',
          '.nav-link',
          '.dropdown-item',
          '.btn'
        ].join(',');
        const controls = Array.from(document.querySelectorAll(controlSelector))
          .filter(visible)
          .slice(0, 500)
          .map((el) => ({
            text: clean(el.innerText || el.textContent),
            tag: el.tagName,
            role: el.getAttribute('role') || '',
            id: el.id || '',
            className: typeof el.className === 'string' ? el.className : '',
            ariaSelected: el.getAttribute('aria-selected') || '',
            ariaPressed: el.getAttribute('aria-pressed') || '',
            disabled: !!el.disabled || el.getAttribute('aria-disabled') === 'true'
          }));

        const activeControls = Array.from(document.querySelectorAll(
          '[aria-selected="true"], [aria-pressed="true"], .active, .selected, .mat-tab-label-active'
        )).filter(visible).map((el) => clean(el.innerText || el.textContent))
          .filter((text) => text && text.length <= 120)
          .slice(0, 100);

        const loadingSelectors = [
          '[aria-busy="true"]',
          '[data-loading="true"]',
          '.spinner-border',
          '.spinner-grow',
          '.mat-progress-spinner',
          'mat-progress-spinner',
          '.ngx-spinner-overlay',
          '.ag-overlay-loading-center',
          '.loading-overlay',
          '.loader'
        ];
        const loadingCount = loadingSelectors.reduce((count, selector) => {
          return count + Array.from(document.querySelectorAll(selector)).filter(visible).length;
        }, 0);

        return {
          title: document.title,
          url: location.href,
          readyState: document.readyState,
          text: clean(document.body ? document.body.innerText : ''),
          tables,
          inputs,
          links,
          controls,
          activeControls,
          loadingCount
        };
        """
    )


def save_debug(company, cin, payload):
    if not SAVE_DEBUG:
        return ""
    os.makedirs(DEBUG_DIR, exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", f"{cin}_{company}")[:120]
    path = os.path.join(DEBUG_DIR, f"{RUN_TIMESTAMP}_{safe}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    return path


def candidate_search_inputs(driver):
    selectors = [
        "input[type='search']",
        "input[placeholder*='Search' i]",
        "input[placeholder*='CIN' i]",
        "input[placeholder*='Company' i]",
        "input[name*='search' i]",
        "input[id*='search' i]",
        "input",
    ]

    seen = set()
    scored_elements = []
    for selector in selectors:
        try:
            for element in driver.find_elements(By.CSS_SELECTOR, selector):
                key = element.id
                if key in seen:
                    continue
                seen.add(key)
                if element.is_displayed() and element.is_enabled():
                    input_type = clean_text(element.get_attribute("type")).lower()
                    if input_type in {
                        "button",
                        "checkbox",
                        "date",
                        "file",
                        "hidden",
                        "number",
                        "radio",
                        "range",
                        "submit",
                    }:
                        continue

                    metadata = " ".join(
                        clean_text(element.get_attribute(attr)).lower()
                        for attr in ["type", "name", "id", "placeholder", "aria-label", "role"]
                    )
                    score = 0
                    if "search" in metadata:
                        score += 8
                    if "cin" in metadata or "llpin" in metadata:
                        score += 6
                    if "company" in metadata:
                        score += 4
                    if input_type == "search":
                        score += 5
                    scored_elements.append((score, element))
        except Exception:
            continue

    scored_elements.sort(key=lambda item: item[0], reverse=True)
    return [element for _, element in scored_elements[:8]]


def listing_result_ready(driver, cin, company):
    return driver.execute_script(
        """
        const cin = arguments[0].toUpperCase();
        const company = arguments[1].toUpperCase();
        const tokens = company.split(/\\s+/).filter((token) => token.length >= 4).slice(0, 4);

        function score(text) {
          const upper = (text || '').toUpperCase();
          let value = upper.includes(cin) ? 10 : 0;
          for (const token of tokens) {
            if (upper.includes(token)) value += 1;
          }
          return value;
        }

        const links = Array.from(document.querySelectorAll('a[href]'));
        for (const link of links) {
          const container = link.closest('tr, [role="row"], .ag-row, .card, .row, li, div') || link;
          const text = `${link.innerText || ''} ${container.innerText || ''} ${link.href || ''}`;
          if (score(text) >= 6 && (link.href || '').includes('/company/')) return true;
        }

        const rows = Array.from(document.querySelectorAll('tr, [role="row"], .ag-row'));
        for (const row of rows) {
          if (score(row.innerText || '') >= 6) return true;
        }

        return false;
        """,
        cin,
        company,
    )


def wait_for_listing_result(driver, cin, company, timeout=SEARCH_RESULT_TIMEOUT_SECONDS):
    end_time = time.time() + timeout
    while time.time() < end_time:
        if listing_result_ready(driver, cin, company):
            return True
        time.sleep(0.75)
    return False


def search_company_by_cin(driver, cin, company):
    for attempt in range(max(1, SEARCH_ATTEMPTS)):
        navigate_to_mca_listing(driver)

        if looks_like_dashboard(driver):
            raise RuntimeError("Still on PrivateCircle dashboard, not Company Master List.")

        inputs = candidate_search_inputs(driver)[:4]
        for input_idx, element in enumerate(inputs):
            try:
                element.click()
                element.send_keys(Keys.CONTROL, "a")
                element.send_keys(Keys.BACKSPACE)
                element.send_keys(cin)
                driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));"
                    "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                    element,
                )
                element.send_keys(Keys.ENTER)

                timeout = SEARCH_RESULT_TIMEOUT_SECONDS if input_idx == 0 else SEARCH_WAIT_SECONDS
                if wait_for_listing_result(driver, cin, company, timeout=max(2, timeout)):
                    return True
            except Exception:
                continue

        if attempt + 1 < max(1, SEARCH_ATTEMPTS):
            time.sleep(1)
    return False


def click_matching_company(driver, cin, company):
    target_url = driver.execute_script(
        """
        const cin = arguments[0].toUpperCase();
        const company = arguments[1].toUpperCase();
        const companyTokens = company.split(/\\s+/).filter((token) => token.length >= 4).slice(0, 4);

        function matchInfo(text) {
          const upper = (text || '').toUpperCase();
          const cinHit = upper.includes(cin);
          let tokenHits = 0;
          for (const token of companyTokens) {
            if (upper.includes(token)) tokenHits += 1;
          }
          return { cinHit, tokenHits };
        }

        function isGoodMatch(info) {
          if (info.cinHit) return true;
          return info.tokenHits >= Math.max(1, Math.min(2, companyTokens.length));
        }

        function profileRank(href) {
          let path = '';
          try {
            path = new URL(href, location.href).pathname.toLowerCase();
          } catch {
            return 0;
          }
          if (path.includes('/company/profile_updated/')) return 60;
          if (path.includes('/company/profile/')) return 50;
          if (path.includes('/company/master_profile_updated/')) return 10;
          return 0;
        }

        function scoreMatch(info, text, href, linkText) {
          const textLength = (text || '').length;
          const ownInfo = matchInfo(`${linkText || ''} ${href || ''}`);
          return (
            (info.cinHit ? 100 : 0)
            + (info.tokenHits * 10)
            + (ownInfo.cinHit ? 25 : 0)
            + (ownInfo.tokenHits * 4)
            + profileRank(href)
            - Math.min(textLength / 1000, 5)
          );
        }

        function isCandidateCompanyHref(href) {
          if (!href) return false;
          let path = '';
          try {
            path = new URL(href, location.href).pathname.toLowerCase();
          } catch {
            return false;
          }
          if (!path.includes('/company/')) return false;
          if (path.includes('/company/mca_new/mca-listing')) return false;
          if (path.includes('/company/company_new/comp-listing')) return false;
          if (path.includes('dashboard') || path.includes('login') || path.includes('logout')) return false;
          return true;
        }

        function closestContainer(link) {
          return link.closest('tr, [role="row"], .ag-row, .card, .row, li') || link;
        }

        let bestHref = null;
        let bestRow = null;
        let bestScore = 0;
        const links = Array.from(document.querySelectorAll('a[href]'));

        for (const link of links) {
          const href = link.href || '';
          if (!isCandidateCompanyHref(href)) continue;
          const container = closestContainer(link);
          const text = `${link.innerText || ''} ${container.innerText || ''} ${link.href || ''}`;
          const info = matchInfo(text);
          if (!isGoodMatch(info)) continue;
          const totalScore = scoreMatch(info, text, href, link.innerText || '');
          if (totalScore > bestScore) {
            bestHref = href;
            bestScore = totalScore;
          }
        }

        if (bestHref) return bestHref;

        const rows = Array.from(document.querySelectorAll('tr, [role="row"], .ag-row'));
        for (const row of rows) {
          const info = matchInfo(row.innerText || '');
          if (!isGoodMatch(info)) continue;
          const score = scoreMatch(info, row.innerText || '', '', '');
          if (score > bestScore) {
            bestRow = row;
            bestScore = score;
          }
        }

        if (!bestRow) return null;
        bestRow.scrollIntoView({ block: 'center' });
        const rowLink = Array.from(bestRow.querySelectorAll('a[href]')).find((link) => isCandidateCompanyHref(link.href || ''));
        if (rowLink) {
          rowLink.scrollIntoView({ block: 'center' });
          rowLink.click();
        } else {
          bestRow.click();
        }
        return "__ROW_CLICKED__";
        """,
        cin,
        company,
    )

    if not target_url:
        return False

    if target_url != "__ROW_CLICKED__":
        driver.get(target_url)

    return wait_for_company_profile(driver, cin, company, timeout=PROFILE_LOAD_TIMEOUT_SECONDS)


def wait_for_company_profile(driver, cin, company, timeout=30):
    end_time = time.time() + timeout
    company_tokens = [token for token in norm(company).split() if len(token) >= 4][:4]

    while time.time() < end_time:
        try:
            url = driver.current_url.lower()
            body = driver.find_element(By.TAG_NAME, "body").text
            body_norm = norm(body)
            cin_ok = norm(cin) in body_norm
            token_hits = sum(1 for token in company_tokens if token in body_norm)
            profile_url = is_company_detail_url(driver.current_url)
            has_profile_content = any(
                label in body_norm
                for label in ["QUICK FACTS", "FINANCIALS", "COMPANY STATUS", "ACTIVE COMPLIANCE"]
            )
            if profile_url and (cin_ok or token_hits >= max(1, min(2, len(company_tokens)))) and has_profile_content:
                return True
        except Exception:
            pass
        time.sleep(0.75)

    return False


def click_text_control(driver, label):
    result = driver.execute_script(
        """
        const clean = (value) => (value || '').replace(/\\s+/g, ' ').trim();
        const target = arguments[0].toUpperCase();
        const selector = [
          'a',
          'button',
          'label',
          '[role="tab"]',
          '[role="button"]',
          '[role="menuitem"]',
          '[role="option"]',
          '[onclick]',
          '[data-toggle="tab"]',
          '[data-bs-toggle="tab"]',
          '.mat-tab-label',
          '.mat-option',
          '.mat-button-toggle-button',
          'mat-button-toggle',
          '.nav-link',
          '.dropdown-item',
          '.btn'
        ].join(',');
        const candidates = Array.from(document.querySelectorAll(selector));
        const matches = candidates.map((el) => {
          if (!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)) return false;
          if (el.disabled || el.getAttribute('aria-disabled') === 'true') return false;
          const text = clean(el.innerText || el.textContent).toUpperCase();
          if (!text) return false;
          if (text.length > Math.max(80, target.length + 40)) return false;
          const exact = text === target;
          if (!exact && !text.startsWith(target + ' ') && !text.endsWith(' ' + target)) return false;

          const role = (el.getAttribute('role') || '').toLowerCase();
          const className = typeof el.className === 'string' ? el.className.toLowerCase() : '';
          let score = exact ? 100 : 60;
          if (role === 'tab') score += 45;
          if (role === 'button' || el.tagName === 'BUTTON') score += 35;
          if (el.tagName === 'A' || el.tagName === 'LABEL') score += 25;
          if (className.includes('mat-tab') || className.includes('nav-link')) score += 30;
          if (className.includes('dropdown') || role === 'option') score += 20;
          if (el.closest('[class*="financial" i], [id*="financial" i], mat-tab-group, .tab-content')) score += 25;
          if (el.closest('table, [role="grid"]')) score -= 120;
          score -= Math.min(text.length, 80) / 100;
          return { el, score, text };
        }).filter(Boolean).sort((a, b) => b.score - a.score);

        const match = matches[0];
        if (!match || match.score < 50) return { clicked: false };
        const node = match.el;
        const wasActive = (
          node.getAttribute('aria-selected') === 'true'
          || node.getAttribute('aria-pressed') === 'true'
          || node.classList.contains('active')
          || node.classList.contains('selected')
          || node.classList.contains('mat-tab-label-active')
        );
        node.scrollIntoView({ block: 'center' });
        if (typeof node.focus === 'function') node.focus({ preventScroll: true });
        node.click();
        return {
          clicked: true,
          tag: node.tagName,
          role: node.getAttribute('role') || '',
          id: node.id || '',
          className: typeof node.className === 'string' ? node.className : '',
          text: match.text,
          wasActive
        };
        """,
        label,
    )
    return result if isinstance(result, dict) else {"clicked": bool(result)}


def click_statement_control(driver, statement_label, control_label):
    """Click a control inside the requested financial-statement panel."""
    result = driver.execute_script(
        """
        const clean = (value) => (value || '').replace(/\\s+/g, ' ').trim();
        const visible = (el) => !!(el && (el.offsetWidth || el.offsetHeight || el.getClientRects().length));
        const statementTarget = clean(arguments[0]).toUpperCase();
        const controlTarget = clean(arguments[1]).toUpperCase();
        const selector = [
          'a',
          'button',
          'label',
          '[role="tab"]',
          '[role="button"]',
          '[data-toggle="tab"]',
          '[data-bs-toggle="tab"]',
          '.mat-tab-label',
          '.mat-button-toggle-button',
          'mat-button-toggle',
          '.nav-link',
          '.btn'
        ].join(',');

        const controls = Array.from(document.querySelectorAll(selector)).filter((el) => {
          if (!visible(el) || el.disabled || el.getAttribute('aria-disabled') === 'true') return false;
          return clean(el.innerText || el.textContent).toUpperCase() === controlTarget;
        });
        if (!controls.length) return { clicked: false, reason: 'Control not found' };

        // PrivateCircle repeats Summary/Detailed and filing-basis controls for
        // each statement. Match by the smallest shared DOM container with the
        // requested statement heading instead of clicking the first label.
        const statementNodeSet = new Set();
        const textWalker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
        let textNode = textWalker.nextNode();
        while (textNode) {
          if (clean(textNode.nodeValue).toUpperCase() === statementTarget && visible(textNode.parentElement)) {
            statementNodeSet.add(textNode.parentElement);
          }
          textNode = textWalker.nextNode();
        }
        const statementNodes = Array.from(statementNodeSet);
        if (!statementNodes.length) return { clicked: false, reason: 'Statement context not found' };

        const ancestors = (el) => {
          const result = [];
          let node = el;
          while (node && node.nodeType === 1) {
            result.push(node);
            node = node.parentElement;
          }
          return result;
        };
        const statementNames = ['INCOME STATEMENT', 'BALANCE SHEET', 'CASH FLOW'];
        let best = null;

        for (const control of controls) {
          const controlAncestors = ancestors(control);
          const controlIndex = new Map(controlAncestors.map((node, index) => [node, index]));
          for (const statementNode of statementNodes) {
            const statementAncestors = ancestors(statementNode);
            let common = null;
            let statementDistance = 0;
            for (let index = 0; index < statementAncestors.length; index += 1) {
              if (controlIndex.has(statementAncestors[index])) {
                common = statementAncestors[index];
                statementDistance = index;
                break;
              }
            }
            if (!common || common === document.body || common === document.documentElement) continue;

            const distance = controlIndex.get(common) + statementDistance;
            const commonText = clean(common.innerText || common.textContent).toUpperCase();
            const otherStatements = statementNames.filter(
              (name) => name !== statementTarget && commonText.includes(name)
            ).length;
            const descendantCount = common.querySelectorAll('*').length;
            const className = typeof control.className === 'string' ? control.className.toLowerCase() : '';
            let score = 2000 - (distance * 120);
            score -= Math.min(descendantCount, 2000) / 2;
            score -= Math.min(commonText.length, 20000) / 200;
            score -= otherStatements * 500;
            if (control.tagName === 'BUTTON') score += 80;
            if ((control.getAttribute('role') || '').toLowerCase() === 'tab') score += 40;
            if (className.includes('mat-button-toggle-button')) score += 30;

            if (!best || score > best.score) {
              best = { control, score, distance };
            }
          }
        }

        if (!best) return { clicked: false, reason: 'No control matched the statement panel' };
        const node = best.control;
        const className = typeof node.className === 'string' ? node.className : '';
        const activeClass = className.toLowerCase();
        const wasActive = (
          node.getAttribute('aria-selected') === 'true'
          || node.getAttribute('aria-pressed') === 'true'
          || /(^|\\s)(active|selected|checked|highlight-tab|tab-top-item1_active)(\\s|$)/.test(activeClass)
          || !!node.closest('.mat-button-toggle-checked, .active, .selected')
        );
        node.scrollIntoView({ block: 'center' });
        if (typeof node.focus === 'function') node.focus({ preventScroll: true });
        node.click();
        return {
          clicked: true,
          tag: node.tagName,
          role: node.getAttribute('role') || '',
          id: node.id || '',
          className,
          text: clean(node.innerText || node.textContent),
          statement: statementTarget,
          wasActive,
          matchDistance: best.distance
        };
        """,
        statement_label,
        control_label,
    )
    return result if isinstance(result, dict) else {"clicked": bool(result)}


def has_financial_statement_content(payload):
    tables = payload.get("tables", [])
    text = payload.get("text", "")
    table_text = " ".join(
        " ".join(" ".join(row) for row in table[:8])
        for table in tables
        if table
    )
    combined = norm(f"{text[:5000]} {table_text}")
    has_statement_labels = any(
        label in combined
        for label in [
            "END OF YEAR",
            "FILING TYPE",
            "REVENUE",
            "EBITDA",
            "BALANCE SHEET",
            "NET FIXED ASSETS",
            "CURRENT LIABILITIES",
        ]
    )
    has_usable_tables = any(
        table and len(table) >= 3 and any(norm(row[0]) in {"END OF YEAR", "FILING TYPE", "REVENUE", "EBITDA"} for row in table if row)
        for table in tables
    )
    return has_statement_labels or has_usable_tables


def extraction_has_metrics(payload, expected_metrics):
    if not expected_metrics:
        return has_financial_statement_content(payload)

    financials = extract_financials(payload)
    return any(clean_text(financials.get(metric)) for metric in expected_metrics)


def payload_has_row_alias(payload, aliases):
    alias_norms = {norm(alias) for alias in aliases}
    for table in payload.get("tables", []):
        for row in table:
            if row and (
                norm(row[0]) in alias_norms
                or financial_row_label(row[0]) in alias_norms
            ):
                return True
    return False


def page_has_no_financial_data(payload):
    text_upper = clean_text(payload.get("text", "")).upper()
    return "NO DATA AVAILABLE" in text_upper and any(
        label in text_upper
        for label in ["FINANCIALS", "INCOME STATEMENT", "BALANCE SHEET"]
    )


def payload_content_signature(payload):
    tables_blob = json.dumps(payload.get("tables", []), ensure_ascii=False, separators=(",", ":"))
    text = clean_text(payload.get("text", ""))
    return (
        clean_text(payload.get("url", "")),
        len(text),
        hash(text),
        hash(tables_blob),
    )


def control_is_active(payload, label):
    target = norm(label)
    if not target:
        return False

    for active_text in payload.get("activeControls", []):
        active = norm(active_text)
        if active == target or active.startswith(f"{target} ") or active.endswith(f" {target}"):
            return True
    return False


def payload_is_busy(payload):
    try:
        return int(payload.get("loadingCount", 0) or 0) > 0
    except (TypeError, ValueError):
        return False


def finish_financial_wait(payload, result, started_at):
    payload["waitResult"] = result
    payload["waitSeconds"] = round(time.time() - started_at, 2)
    return payload


def wait_until_financial_data_extracted(
    driver,
    expected_metrics=None,
    expected_row_aliases=None,
    timeout=FINANCIAL_LOAD_TIMEOUT_SECONDS,
    previous_signature=None,
    clicked_label=None,
    label_was_active=False,
):
    """Wait for the clicked state to settle and expose extractable statement data."""
    started_at = time.time()
    end_time = started_at + timeout
    last_payload = page_payload(driver)
    stable_seen = 0
    last_signature = payload_content_signature(last_payload)
    transition_seen = previous_signature is None or label_was_active
    busy_seen = payload_is_busy(last_payload)

    while time.time() < end_time:
        last_payload = page_payload(driver)
        signature = payload_content_signature(last_payload)
        if signature == last_signature:
            stable_seen += 1
        else:
            stable_seen = 0
            last_signature = signature

        if previous_signature is not None and signature != previous_signature:
            transition_seen = True

        busy = payload_is_busy(last_payload)
        if busy:
            busy_seen = True
        elif busy_seen:
            transition_seen = True

        elapsed = time.time() - started_at
        active_now = clicked_label and control_is_active(last_payload, clicked_label)
        if (
            not transition_seen
            and active_now
            and elapsed >= FINANCIAL_UNCHANGED_GRACE_SECONDS
            and stable_seen >= 2
        ):
            # Some PrivateCircle controls update only CSS state when the selected
            # view has the same values. Give the request time to settle first.
            transition_seen = True

        if (
            not transition_seen
            and elapsed >= max(10, FINANCIAL_UNCHANGED_GRACE_SECONDS * 2)
            and stable_seen >= 4
        ):
            # A click can be valid even when the site exposes no active-state or
            # loading marker. Do not spend the full safety timeout on a no-op view.
            transition_seen = True

        settled = transition_seen and not busy and elapsed >= 1 and stable_seen >= 1
        if settled:
            if expected_row_aliases and payload_has_row_alias(last_payload, expected_row_aliases):
                return finish_financial_wait(last_payload, "Expected statement row detected", started_at)
            if extraction_has_metrics(last_payload, expected_metrics):
                return finish_financial_wait(last_payload, "Expected metric detected", started_at)
            if page_has_no_financial_data(last_payload):
                return finish_financial_wait(last_payload, "No financial data shown", started_at)

            # The requested row may genuinely be absent. Once the statement has
            # stopped changing, keep the available table and let extraction report it.
            if stable_seen >= 4 and has_financial_statement_content(last_payload):
                return finish_financial_wait(last_payload, "Statement settled without requested row", started_at)

        time.sleep(0.5)

    return finish_financial_wait(last_payload, "Safety timeout reached", started_at)


def merge_payloads(payloads):
    if not payloads:
        return {}

    merged = dict(payloads[-1])
    merged["text"] = " ".join(clean_text(payload.get("text", "")) for payload in payloads if payload.get("text"))
    merged["tables"] = []
    merged["links"] = []
    merged["inputs"] = []
    merged["captureLog"] = []

    seen_tables = set()
    for payload in payloads:
        for table in payload.get("tables", []):
            key = json.dumps(table, ensure_ascii=False)
            if key not in seen_tables:
                merged["tables"].append(table)
                seen_tables.add(key)
        merged["links"].extend(payload.get("links", []))
        merged["inputs"].extend(payload.get("inputs", []))
        merged["captureLog"].extend(payload.get("captureLog", []))

    return merged


def capture_after_optional_click(
    driver,
    label,
    payloads,
    expected_metrics=None,
    expected_row_aliases=None,
    timeout=None,
    statement_context=None,
):
    before_payload = page_payload(driver)
    click_info = {"clicked": False}
    try:
        if not label:
            click_info = {"clicked": True, "wasActive": True}
        elif statement_context:
            click_info = click_statement_control(driver, statement_context, label)
        else:
            click_info = click_text_control(driver, label)
    except Exception as exc:
        click_info = {"clicked": False, "error": clean_text(exc)}

    clicked = bool(click_info.get("clicked"))
    if clicked:
        payload = wait_until_financial_data_extracted(
            driver,
            expected_metrics=expected_metrics,
            expected_row_aliases=expected_row_aliases,
            timeout=timeout or FINANCIAL_LOAD_TIMEOUT_SECONDS,
            previous_signature=payload_content_signature(before_payload) if label else None,
            clicked_label=label,
            label_was_active=bool(click_info.get("wasActive")),
        )
    else:
        payload = before_payload

    payload.setdefault("captureLog", []).append(
        {
            "label": label or "Initial state",
            "statementContext": statement_context or "",
            "clicked": clicked,
            "matchedTag": clean_text(click_info.get("tag", "")),
            "matchedRole": clean_text(click_info.get("role", "")),
            "matchedId": clean_text(click_info.get("id", "")),
            "matchedClass": clean_text(click_info.get("className", "")),
            "matchedText": clean_text(click_info.get("text", "")),
            "matchDistance": click_info.get("matchDistance", ""),
            "wasActive": bool(click_info.get("wasActive")),
            "waitResult": clean_text(payload.get("waitResult", "")),
            "waitSeconds": payload.get("waitSeconds", ""),
            "error": clean_text(click_info.get("error") or click_info.get("reason", "")),
        }
    )
    payloads.append(payload)
    return clicked


def collect_company_payload(driver, sections=None):
    requested_sections = set(sections or ["income", "balance"])
    initial_payload = wait_until_financial_data_extracted(driver, timeout=8)
    initial_payload.setdefault("captureLog", []).append(
        {"label": "Profile overview", "clicked": False}
    )
    payloads = [initial_payload]

    for label in ["View Financials", "Financials"]:
        if capture_after_optional_click(driver, label, payloads):
            break

    for view in ["Yearly", "Detailed"]:
        capture_after_optional_click(
            driver,
            view,
            payloads,
            timeout=FINANCIAL_LOAD_TIMEOUT_SECONDS if view == "Detailed" else 15,
        )

    balance_borrowing_aliases = [
        "Long Term Borrowings",
        "Short Term Borrowings",
        "Non Current Borrowings",
        "Current Borrowings",
        "Borrowings",
        "Total Borrowings",
    ]
    for section, statement, expected_metrics, expected_row_aliases in [
        ("income", "Income Statement", ["Revenue", "EBITDA"], None),
        (
            "balance",
            "Balance Sheet",
            ["Long Term Debt", "Short Term Debt"],
            balance_borrowing_aliases,
        ),
    ]:
        if section not in requested_sections:
            continue
        capture_after_optional_click(driver, statement, payloads, expected_metrics, expected_row_aliases)

        for view in ["Yearly", "Detailed"]:
            capture_after_optional_click(
                driver,
                view,
                payloads,
                expected_metrics,
                expected_row_aliases,
                timeout=FINANCIAL_LOAD_TIMEOUT_SECONDS if view == "Detailed" else 15,
                statement_context=statement,
            )

        # Capture standalone as a fallback, then consolidated so the preferred
        # basis remains selected. Each click is stored as a separate page state.
        for basis in ["Standalone", "Consolidated"]:
            capture_after_optional_click(
                driver,
                basis,
                payloads,
                expected_metrics,
                expected_row_aliases,
                statement_context=statement,
            )

    return merge_payloads(payloads)


def flatten_tables(tables):
    rows = []
    for table_idx, table in enumerate(tables or []):
        for row_idx, row in enumerate(table):
            rows.append(
                {
                    "table_idx": table_idx,
                    "row_idx": row_idx,
                    "cells": [clean_text(cell) for cell in row],
                    "text": " | ".join(clean_text(cell) for cell in row if clean_text(cell)),
                }
            )
    return rows


def find_labeled_value(payload, labels):
    text = payload.get("text", "")
    rows = flatten_tables(payload.get("tables", []))
    label_norms = [norm(label) for label in labels]

    for row in rows:
        cells = row["cells"]
        for idx, cell in enumerate(cells):
            cell_norm = norm(cell)
            if any(label in cell_norm for label in label_norms):
                for next_cell in cells[idx + 1 :]:
                    if clean_text(next_cell):
                        return clean_text(next_cell)
                if len(cells) == 2 and clean_text(cells[1]):
                    return clean_text(cells[1])

    for label in labels:
        pattern = re.compile(
            rf"{re.escape(label)}\\s*[:\\-]?\\s*([^\\n\\r|]+)",
            re.IGNORECASE,
        )
        match = pattern.search(text)
        if match:
            return clean_text(match.group(1))
    return ""


def extract_text_between(text, label, stop_labels):
    compact = clean_text(text)
    upper = compact.upper()
    label_upper = label.upper()
    start = upper.find(label_upper)
    if start < 0:
        return ""

    value_start = start + len(label_upper)
    value_end = len(compact)
    for stop_label in stop_labels:
        stop_upper = stop_label.upper()
        stop_at = upper.find(stop_upper, value_start)
        if stop_at >= 0:
            value_end = min(value_end, stop_at)

    return clean_text(compact[value_start:value_end].strip(" :-|"))


def extract_table_value_by_header(payload, header_label):
    label_norm = norm(header_label)
    tables = payload.get("tables", [])

    for idx, table in enumerate(tables):
        if not table:
            continue

        header = [clean_text(cell) for cell in table[0]]
        header_norms = [norm(cell) for cell in header]
        if label_norm not in header_norms:
            continue

        header_idx = header_norms.index(label_norm)

        # PrivateCircle sometimes exposes the header and the result row as separate tables.
        candidate_rows = table[1:]
        if not candidate_rows and idx + 1 < len(tables):
            candidate_rows = tables[idx + 1]

        for row in candidate_rows:
            if header_idx < len(row) and clean_text(row[header_idx]):
                return clean_text(row[header_idx])

    return ""


def normalize_status_value(value):
    cleaned = clean_text(value)
    if norm(cleaned) in {"", "STATUS", "COMPANY STATUS", "CURRENT STATUS", "CIRP STATUS", "STATUS UNDER CIRP"}:
        return ""
    return cleaned


def extract_status(payload):
    text = payload.get("text", "")
    quick_status = extract_text_between(
        text,
        "Quick Facts Status",
        ["Active Compliance", "Status Under CIRP", "Authorized Capital", "Paid Up Capital", "Address"],
    )
    active_compliance = extract_text_between(
        text,
        "Active Compliance",
        ["Status Under CIRP", "Authorized Capital", "Paid Up Capital", "Address", "Company Class"],
    )
    active_idx = text.upper().find("ACTIVE COMPLIANCE")
    status_scope = text[active_idx:] if active_idx >= 0 else text
    cirp_status = extract_text_between(
        status_scope,
        "Status Under CIRP",
        ["Authorized Capital", "Paid Up Capital", "Address", "Company Class", "Subcategory", "Listing Status"],
    )

    company_status = find_labeled_value(
        payload,
        ["Company Status", "CompanyStatus", "MCA Status", "Current Status"],
    )
    if not company_status:
        company_status = quick_status
    if not company_status:
        company_status = extract_table_value_by_header(payload, "Status")

    if not cirp_status:
        cirp_status = find_labeled_value(
            payload,
            ["CIRP Status", "IBC Status", "Insolvency Status", "Process Status"],
        )

    company_status = normalize_status_value(company_status)
    cirp_status = normalize_status_value(cirp_status)

    if active_compliance and "ACTIVE" in active_compliance.upper() and "INACTIVE" not in active_compliance.upper():
        company_status = "Active"

    if not cirp_status:
        exact_cirp_match = re.search(
            r"(?:CIRP Status|IBC Status|Status Under CIRP)\s*[:\-]?\s*(Under CIRP|CIRP|Under Liquidation|Liquidation)",
            text,
            re.IGNORECASE,
        )
        if exact_cirp_match:
            cirp_status = clean_text(exact_cirp_match.group(1))

    if not company_status and re.search(r"\bACTIVE\b", text, re.IGNORECASE):
        company_status = "Active"

    if cirp_status.upper() == "CIRP":
        cirp_status = "Under CIRP"

    return company_status, cirp_status


def is_active_under_cirp(company_status, cirp_status, payload):
    status_upper = clean_text(company_status).upper()
    cirp_upper = clean_text(cirp_status).upper()

    is_active = "ACTIVE" in status_upper and not any(
        bad in status_upper
        for bad in ["INACTIVE", "STRIKE", "STRUCK", "DISSOLVED", "AMALGAMATED", "LIQUIDATION"]
    )
    is_cirp = (
        bool(re.search(r"\bUNDER\s+CIRP\b", cirp_upper))
        or cirp_upper == "CIRP"
        or "CORPORATE INSOLVENCY RESOLUTION PROCESS" in cirp_upper
    )
    not_liquidation = "LIQUIDATION" not in cirp_upper and "LIQUIDATED" not in cirp_upper
    return is_active and is_cirp and not_liquidation


def extract_number(value):
    raw = clean_text(value)
    if not raw or raw in {"-", "--", "nm", "NM"}:
        return None

    match = re.search(
        r"-?(?:[0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)(?:\.[0-9]+)?",
        raw,
    )
    if not match:
        return None

    number = float(match.group(0).replace(",", ""))
    if "(" in raw[: match.start() + 1] and ")" in raw[match.end() :]:
        number = -abs(number)
    # Statement columns already share one displayed unit, so preserve that scale.
    return number


def add_numbers(left, right):
    left_num = extract_number(left)
    right_num = extract_number(right)
    if left_num is None or right_num is None:
        return ""
    total = (left_num or 0) + (right_num or 0)
    return round(total, 2)


def normalized_row_label(row):
    if not row:
        return ""
    return norm(row[0])


def year_from_cell(cell):
    match = re.search(r"\b(19|20)\d{2}\b", clean_text(cell))
    if not match:
        return None
    return int(match.group(0))


def is_quarterly_table(table):
    for row in table:
        if not row:
            continue
        if norm(row[0]) == "PERIOD":
            period_values = " ".join(clean_text(cell).upper() for cell in row[1:])
            if re.search(r"\bQ[1-4]\b", period_values):
                return True
    return False


def annual_financial_tables(payload):
    annual_tables = []
    for table_idx, table in enumerate(payload.get("tables", [])):
        if not table or is_quarterly_table(table):
            continue

        year_row_idx = None
        for idx, row in enumerate(table):
            if row and norm(row[0]) in {"END OF YEAR", "END YEAR", "YEAR"}:
                year_row_idx = idx
                break
        if year_row_idx is None:
            continue

        year_row = table[year_row_idx]
        year_cols = {}
        for col_idx, cell in enumerate(year_row[1:], start=1):
            year = year_from_cell(cell)
            if year:
                year_cols[col_idx] = year
        if not year_cols:
            continue

        filing_types = {}
        for row in table:
            if row and norm(row[0]) == "FILING TYPE":
                for col_idx, cell in enumerate(row[1:], start=1):
                    filing_types[col_idx] = clean_text(cell).title()
                break

        annual_tables.append(
            {
                "table_idx": table_idx,
                "rows": table,
                "year_cols": year_cols,
                "filing_types": filing_types,
            }
        )

    return annual_tables


def row_matches_metric(row_label, aliases):
    label = financial_row_label(row_label)
    alias_norms = {financial_row_label(alias) for alias in aliases}
    return label in alias_norms


NON_CURRENT_LIABILITY_HEADINGS = {
    "NON CURRENT LIABILITIES",
    "NONCURRENT LIABILITIES",
}
CURRENT_LIABILITY_HEADINGS = {
    "CURRENT LIABILITIES",
}
LIABILITY_SECTION_BOUNDARIES = {
    "ASSETS",
    "CURRENT ASSETS",
    "EQUITY",
    "EQUITY AND LIABILITIES",
    "NET WORKING CAPITAL",
    "NON CURRENT ASSETS",
    "NONCURRENT ASSETS",
    "SHAREHOLDERS FUNDS",
    "TOTAL APPLICATIONS",
    "TOTAL ASSETS",
    "TOTAL EQUITY AND LIABILITIES",
    "TOTAL LIABILITIES",
    "TOTAL SOURCES",
}
LONG_TERM_BORROWING_LABELS = {
    "LONG TERM BORROWINGS",
    "NON CURRENT BORROWINGS",
    "NONCURRENT BORROWINGS",
}
SHORT_TERM_BORROWING_LABELS = {
    "CURRENT BORROWINGS",
    "SHORT TERM BORROWINGS",
}
GENERIC_BORROWING_LABELS = {
    "BORROWINGS",
    "TOTAL BORROWINGS",
}
OUTLINE_PREFIXES = {
    "I",
    "II",
    "III",
    "IV",
    "V",
    "VI",
    "VII",
    "VIII",
    "IX",
    "X",
}


def financial_row_label(row_label):
    """Normalize a statement row and remove outline prefixes such as '(i)' or 'a)'."""
    tokens = norm(row_label).split()
    while tokens and (
        tokens[0].isdigit()
        or tokens[0] in OUTLINE_PREFIXES
        or (len(tokens[0]) == 1 and tokens[0].isalpha())
    ):
        tokens.pop(0)
    return " ".join(tokens)


def row_matches_borrowing_label(label, aliases):
    if label in aliases:
        return True
    if label.startswith("TOTAL ") and label[6:] in aliases:
        return True
    return any(label.startswith(f"{alias} ") for alias in aliases)


def canonical_financial_basis(basis):
    basis_upper = clean_text(basis).upper()
    if "CONSOLIDATED" in basis_upper:
        return "Consolidated"
    if "STANDALONE" in basis_upper:
        return "Standalone"
    return "Unknown"


def empty_metric(**extra):
    return {
        "value": "",
        "year": None,
        "basis": "",
        "metric_label": "",
        **extra,
    }


def metric_candidates_from_annual_tables(payload, aliases):
    candidates = []
    for table in annual_financial_tables(payload):
        for row in table["rows"]:
            if not row or not row_matches_metric(row[0], aliases):
                continue

            for col_idx, year in table["year_cols"].items():
                if col_idx >= len(row):
                    continue
                value = clean_text(row[col_idx])
                if extract_number(value) is None:
                    continue
                basis = table["filing_types"].get(col_idx, "Unknown")
                candidates.append(
                    {
                        "value": value,
                        "year": year,
                        "basis": basis,
                        "table_idx": table["table_idx"],
                        "metric_label": clean_text(row[0]),
                    }
                )

    return candidates


def borrowing_candidates_from_annual_tables(payload, liability_section):
    """Extract borrowings from the requested current/non-current liability section."""
    if liability_section not in {"current", "non_current"}:
        raise ValueError(f"Unsupported liability section: {liability_section}")

    explicit_labels = (
        LONG_TERM_BORROWING_LABELS
        if liability_section == "non_current"
        else SHORT_TERM_BORROWING_LABELS
    )
    candidates = []

    for table in annual_financial_tables(payload):
        current_section = ""
        for row in table["rows"]:
            if not row:
                continue

            label = financial_row_label(row[0])
            if label in NON_CURRENT_LIABILITY_HEADINGS:
                current_section = "non_current"
                continue
            if label in CURRENT_LIABILITY_HEADINGS:
                current_section = "current"
                continue
            if label in LIABILITY_SECTION_BOUNDARIES:
                current_section = ""

            explicit_match = row_matches_borrowing_label(label, explicit_labels)
            generic_match = row_matches_borrowing_label(label, GENERIC_BORROWING_LABELS)
            if explicit_match:
                # A directional label can stand on its own, but never accept it
                # when it appears under the opposite liability section.
                if current_section and current_section != liability_section:
                    continue
                label_rank = 2
            elif generic_match and current_section == liability_section:
                label_rank = 1
            else:
                continue

            for col_idx, year in table["year_cols"].items():
                if col_idx >= len(row):
                    continue
                value = clean_text(row[col_idx])
                if extract_number(value) is None:
                    continue
                basis = table["filing_types"].get(col_idx, "Unknown")
                candidates.append(
                    {
                        "value": value,
                        "year": year,
                        "basis": basis,
                        "table_idx": table["table_idx"],
                        "metric_label": clean_text(row[0]),
                        "liability_section": liability_section,
                        "label_rank": label_rank,
                    }
                )

    return candidates


def latest_financial_snapshot(candidate_sets):
    candidates = [candidate for values in candidate_sets.values() for candidate in values]
    years = [candidate.get("year") for candidate in candidates if candidate.get("year")]
    if not years:
        return None, ""

    latest_year = max(years)
    latest_year_bases = {
        canonical_financial_basis(candidate.get("basis"))
        for candidate in candidates
        if candidate.get("year") == latest_year
    }
    if "Consolidated" in latest_year_bases:
        basis = "Consolidated"
    elif "Standalone" in latest_year_bases:
        basis = "Standalone"
    else:
        basis = "Unknown"
    return latest_year, basis


def select_snapshot_candidate(candidates, target_year, target_basis, **empty_extra):
    matching = [
        candidate
        for candidate in candidates
        if candidate.get("year") == target_year
        and canonical_financial_basis(candidate.get("basis")) == target_basis
    ]
    if not matching:
        return empty_metric(**empty_extra)

    matching.sort(
        key=lambda item: (item.get("label_rank", 0), -item.get("table_idx", 0)),
        reverse=True,
    )
    return matching[0]


def financial_notes_for_metrics(metrics, financial_year=None, financial_basis=""):
    notes = []
    for label, data in metrics.items():
        if not data.get("value"):
            missing_description = {
                "Long Term Debt": "Long-term borrowings under Non-Current Liabilities",
                "Short Term Debt": "Short-term borrowings under Current Liabilities",
            }.get(label, label)
            notes.append(f"{missing_description} not found in annual financial tables.")

    if financial_basis and financial_basis != "Consolidated":
        notes.append("Consolidated annual values were not detected; using standalone/available annual filing basis.")
    if financial_year:
        notes.append(
            f"All reported metrics are restricted to FY {financial_year}; older years were not used to fill missing values."
        )

    return notes


def extract_financial_units(payload):
    text = clean_text(payload.get("text", ""))
    matches = re.findall(r"\bALL\s+VALUES\s+IN\s+(.{1,16}?)\s+FROM\b", text, re.IGNORECASE)
    units = []
    for match in matches:
        unit = clean_text(match).strip(" ,;:|")
        if unit and unit not in units:
            units.append(unit)
    return units[0] if len(units) == 1 else ("Mixed" if units else "")


def financial_source_note(metrics, financial_units):
    sources = []
    for label, data in metrics.items():
        if not data.get("value"):
            continue
        row_label = clean_text(data.get("metric_label")) or label
        basis = clean_text(data.get("basis")) or "Unknown basis"
        year = data.get("year") or "Unknown year"
        sources.append(f"{label} <- {row_label} [{basis}, FY {year}]")

    if not sources:
        return ""

    note = "PrivateCircle statement sources: " + "; ".join(sources) + "."
    if financial_units:
        note += f" Statement units: {financial_units}."
    return note


def extract_financials(payload):
    candidate_sets = {
        "Revenue": metric_candidates_from_annual_tables(
            payload,
            ["Revenue", "Total Revenue", "Revenue From Operations", "Operating Revenue", "Sales"],
        ),
        "EBITDA": metric_candidates_from_annual_tables(payload, ["EBITDA"]),
        "Long Term Debt": borrowing_candidates_from_annual_tables(payload, "non_current"),
        "Short Term Debt": borrowing_candidates_from_annual_tables(payload, "current"),
        "Inventory": metric_candidates_from_annual_tables(payload, ["Inventory", "Inventories"]),
        "Fixed Assets": metric_candidates_from_annual_tables(
            payload,
            [
                "Net Fixed Assets",
                "Total Net Fixed Assets",
                "Fixed Assets",
                "Fixed Assets Non Current",
                "Property Plant And Equipment",
                "Net Property Plant And Equipment",
                "PPE",
                "Tangible Assets",
                "Tangible Assets Net Block",
            ],
        ),
    }

    financial_year, financial_basis = latest_financial_snapshot(candidate_sets)
    metrics = {}
    for label, candidates in candidate_sets.items():
        empty_extra = {}
        if label == "Long Term Debt":
            empty_extra["liability_section"] = "non_current"
        elif label == "Short Term Debt":
            empty_extra["liability_section"] = "current"
        metrics[label] = select_snapshot_candidate(
            candidates,
            financial_year,
            financial_basis,
            **empty_extra,
        )

    values = {label: data.get("value", "") for label, data in metrics.items()}
    financial_units = extract_financial_units(payload)
    notes = financial_notes_for_metrics(metrics, financial_year, financial_basis)
    source_note = financial_source_note(metrics, financial_units)
    if source_note:
        notes.append(source_note)

    return {
        "Financial Basis": financial_basis,
        "Financial Year": financial_year,
        "Financial Units": financial_units,
        "Revenue": values["Revenue"],
        "EBITDA": values["EBITDA"],
        "Long Term Debt": values["Long Term Debt"],
        "Short Term Debt": values["Short Term Debt"],
        "Debt (LT + ST)": add_numbers(values["Long Term Debt"], values["Short Term Debt"]),
        "Inventory": values["Inventory"],
        "Fixed Assets": values["Fixed Assets"],
        "Asset (Inventory + Fixed)": add_numbers(values["Inventory"], values["Fixed Assets"]),
        "_financial_notes": " ".join(notes),
    }


def missing_financial_sections(financials):
    sections = []
    if not financials.get("Revenue") or not financials.get("EBITDA"):
        sections.append("income")
    if any(
        not financials.get(metric)
        for metric in ["Long Term Debt", "Short Term Debt", "Fixed Assets"]
    ):
        sections.append("balance")
    return sections


def extract_company_data(driver, input_row):
    company = input_row["Corporate Debtor"]
    cin = input_row["CIN / LLPIN"]

    search_ok = search_company_by_cin(driver, cin, company)
    if not search_ok:
        payload = page_payload(driver)
        save_debug(company, cin, payload)
        return {
            **input_row,
            "Extraction Status": "Search failed",
            "Notes": "No visible result for CIN.",
            "PrivateCircle URL": safe_current_url(driver),
        }

    opened_profile = click_matching_company(driver, cin, company)
    payload = collect_company_payload(driver) if opened_profile else page_payload(driver)

    if opened_profile and is_full_company_profile_url(payload.get("url", safe_current_url(driver))):
        for capture_attempt in range(1, max(1, FINANCIAL_CAPTURE_ATTEMPTS)):
            preview_financials = extract_financials(payload)
            missing_sections = missing_financial_sections(preview_financials)
            if not missing_sections:
                break

            print(
                "    -> Retrying slow/missing financial section(s): "
                + ", ".join(missing_sections)
            )
            retry_payload = collect_company_payload(driver, sections=missing_sections)
            payload = merge_payloads([payload, retry_payload])

    save_debug(company, cin, payload)

    company_status, cirp_status = extract_status(payload)
    financials = extract_financials(payload)
    financial_notes = financials.pop("_financial_notes", "")

    active_under_cirp = is_active_under_cirp(company_status, cirp_status, payload)

    notes = []
    if not opened_profile:
        notes.append("Could not confidently open the company profile link from search results.")
    elif is_master_company_profile_url(payload.get("url", safe_current_url(driver))):
        notes.append(
            "PrivateCircle exposed only the MCA master profile; a full financial profile link was not available."
        )
    if financial_notes:
        notes.append(financial_notes)
    if not financials.get("Revenue"):
        notes.append("Revenue not detected.")
    if not financials.get("EBITDA"):
        notes.append("EBITDA not detected.")

    return {
        **input_row,
        "Company Status": company_status,
        "CIRP Status": cirp_status,
        "Active Under CIRP": "Yes" if active_under_cirp else "No",
        **financials,
        "PrivateCircle URL": payload.get("url", safe_current_url(driver)),
        "Extraction Status": "Checked",
        "Notes": " ".join(notes),
    }


def format_output(file_path):
    wb = openpyxl.load_workbook(file_path)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                value = clean_text(cell.value)
                if "\n" in value:
                    value = max(value.split("\n"), key=len)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)

    wb.save(file_path)


def write_output(rows, checkpoint=False):
    suffix = "_checkpoint" if checkpoint else ""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_name = os.path.join(
        OUTPUT_DIR,
        f"PrivateCircle_Company_Debt_Profile_{RUN_TIMESTAMP}{suffix}.xlsx",
    )

    all_df = pd.DataFrame(rows)
    for column in OUTPUT_COLUMNS:
        if column not in all_df.columns:
            all_df[column] = ""

    all_df = all_df[OUTPUT_COLUMNS]
    all_df["Notes"] = all_df["Notes"].map(redact_local_paths)
    active_cirp_df = all_df[all_df["Active Under CIRP"] == "Yes"].copy()

    with pd.ExcelWriter(output_name, engine="openpyxl") as writer:
        all_df.to_excel(writer, index=False, sheet_name="All Companies")
        active_cirp_df.to_excel(writer, index=False, sheet_name="Active CIRP View")

    format_output(output_name)
    return output_name, len(all_df), len(active_cirp_df)


def run_privatecircle_enrichment():
    print("==============================================")
    print("   PRIVATECIRCLE COMPANY DEBT PROFILE ENRICHER")
    print("==============================================\n")

    cin_file = get_latest_cin_file()
    if not cin_file:
        print("[!] No CIN_Directory_*.xlsx file found. Run cin_enricher.py first.")
        return

    print(f"[*] Reading CIN directory: {cin_file}")
    cin_rows = load_cin_rows(cin_file)
    print(f"[*] CINs to check: {len(cin_rows)}")
    if not cin_rows:
        print("[!] No usable CIN/LLPIN rows found.")
        return

    driver = setup_driver()
    results = []
    checkpoint_name = None

    try:
        wait_for_manual_login(driver)

        for idx, row in enumerate(cin_rows, start=1):
            print(f"[{idx}/{len(cin_rows)}] Checking {row['Corporate Debtor'][:70]} ({row['CIN / LLPIN']})")
            result = None
            last_error = None
            for company_attempt in range(max(1, COMPANY_ATTEMPTS)):
                try:
                    result = extract_company_data(driver, row)
                    break
                except Exception as exc:
                    last_error = exc
                    if company_attempt + 1 >= max(1, COMPANY_ATTEMPTS):
                        break

                    print(
                        f"    -> Selenium navigation failed; retrying company "
                        f"({company_attempt + 2}/{max(1, COMPANY_ATTEMPTS)})."
                    )
                    try:
                        if page_needs_login(driver):
                            wait_for_manual_login(driver)
                        else:
                            navigate_to_mca_listing(driver)
                    except Exception:
                        pass

            if result is None:
                result = {
                    **row,
                    "Extraction Status": "Error",
                    "PrivateCircle URL": safe_current_url(driver),
                    "Notes": clean_text(last_error),
                }
            print(f"    -> {result.get('Extraction Status', '')}: {result.get('Company Status', '')} / {result.get('CIRP Status', '')}")
            results.append(result)
            if CHECKPOINT_EVERY > 0 and idx % CHECKPOINT_EVERY == 0:
                checkpoint_name, _, _ = write_output(results, checkpoint=True)
                print(f"    -> Checkpoint saved: {checkpoint_name}")

        output_name, checked_count, active_cirp_count = write_output(results)
        if checkpoint_name and os.path.exists(checkpoint_name):
            os.remove(checkpoint_name)
        print("\n==============================================")
        print("PRIVATECIRCLE ENRICHMENT COMPLETE")
        print("==============================================")
        print(f"Companies checked       : {checked_count}")
        print(f"Active + under CIRP     : {active_cirp_count}")
        print(f"Output workbook         : {output_name}")
        if SAVE_DEBUG:
            print(f"Debug snapshots folder  : {DEBUG_DIR}")

    finally:
        keep_open = os.getenv("PRIVATECIRCLE_KEEP_BROWSER_OPEN", "0") == "1"
        if not keep_open:
            driver.quit()
        else:
            print("[INFO] Chrome is being left open. Set PRIVATECIRCLE_KEEP_BROWSER_OPEN=0 to auto-close it.")


if __name__ == "__main__":
    run_privatecircle_enrichment()
