import glob
import os
import re
import shutil
import tempfile
import time
import urllib.parse
from datetime import date, datetime

import openpyxl
import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill

from nclt_utils import (
    configured_output_directory,
    extract_case_id_from_cells,
    recover_case_id_from_row,
    repair_doubled_ocr,
)


PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DOWNLOAD_DIR = os.path.join(PROJECT_DIR, "downloads")
RUN_TIMESTAMP = os.getenv("NCLT_RUN_TIMESTAMP", "").strip() or datetime.now().strftime(
    "%Y-%m-%d_%H-%M-%S"
)
OUTPUT_DIR = configured_output_directory()
RUN_DOWNLOAD_DIR = os.path.join(BASE_DOWNLOAD_DIR, RUN_TIMESTAMP)
KEEP_BROWSER_LOGS = os.getenv("NCLT_KEEP_BROWSER_LOGS", "0").strip().lower() in {
    "1",
    "true",
    "yes",
}
BASE_URL = "https://nclt.gov.in/index.php/all-cause-list"
COURT_OPEN_MARKER = "\u25ba\u25ba\u25ba"
COURT_CLOSE_MARKER = "\u25c4\u25c4\u25c4"
NEXT_PAGE_SELECTORS = (
    "nav.pager li.pager__item--next:not(.is-disabled) a[href]",
    "nav.pager a[rel='next'][href]",
    "nav.pager a[title*='next'][href]",
    "nav[aria-label*='agination'] a[rel='next'][href]",
    "ul.pagination li.next:not(.disabled) a[href]",
)

LINE_TABLE_SETTINGS = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
}
TEXT_TABLE_SETTINGS = {
    "vertical_strategy": "text",
    "horizontal_strategy": "text",
    "intersection_x_tolerance": 20,
    "snap_tolerance": 5,
}


def build_chrome_options(profile_dir):
    from selenium.webdriver.chrome.options import Options

    chrome_options = Options()
    prefs = {
        "download.default_directory": RUN_DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
    }
    chrome_options.add_experimental_option("prefs", prefs)
    chrome_options.add_argument(f"--user-data-dir={profile_dir}")
    chrome_options.add_argument("--remote-debugging-port=0")
    chrome_options.add_argument("--no-first-run")
    chrome_options.add_argument("--no-default-browser-check")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    return chrome_options


def find_cached_chromedriver(browser_version):
    if not browser_version:
        return None
    version_pattern = (
        browser_version
        if browser_version.count(".") >= 3
        else f"{browser_version}*"
    )
    cache_patterns = (
        os.path.join(
            os.path.expanduser("~"),
            ".wdm",
            "drivers",
            "chromedriver",
            "*",
            version_pattern,
            "**",
            "chromedriver.exe",
        ),
        os.path.join(
            os.path.expanduser("~"),
            ".cache",
            "selenium",
            "chromedriver",
            "*",
            version_pattern,
            "chromedriver.exe",
        ),
    )
    matches = []
    for pattern in cache_patterns:
        matches.extend(glob.glob(pattern, recursive=True))
    matches = [path for path in matches if os.path.isfile(path)]
    return max(matches, key=os.path.getmtime) if matches else None


def setup_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    from webdriver_manager.core.os_manager import ChromeType, OperationSystemManager

    os.makedirs(RUN_DOWNLOAD_DIR, exist_ok=True)
    browser_version = OperationSystemManager().get_browser_version_from_os(
        ChromeType.GOOGLE
    )
    if browser_version:
        print(f"[BROWSER] Installed Chrome: {browser_version}")

    driver_sources = []
    cached_driver = find_cached_chromedriver(browser_version)
    if cached_driver:
        driver_sources.append(
            ("cached Chrome build match", lambda path=cached_driver: path)
        )
    elif browser_version:
        driver_sources.append(
            (
                "Chrome build match",
                lambda: ChromeDriverManager(
                    driver_version=browser_version
                ).install(),
            )
        )
    driver_sources.extend(
        [
            ("latest compatible cached driver", lambda: ChromeDriverManager().install()),
            ("Selenium Manager", lambda: None),
        ]
    )

    errors = []
    attempted_paths = set()
    for attempt, (source_name, resolve_driver_path) in enumerate(driver_sources, 1):
        profile_dir = tempfile.mkdtemp(
            prefix=f".chrome_profile_{attempt}_",
            dir=RUN_DOWNLOAD_DIR,
        )
        log_path = os.path.join(
            RUN_DOWNLOAD_DIR,
            f"chromedriver_startup_{attempt}.log",
        )
        try:
            driver_path = resolve_driver_path()
            normalized_path = os.path.normcase(driver_path) if driver_path else None
            if normalized_path and normalized_path in attempted_paths:
                shutil.rmtree(profile_dir, ignore_errors=True)
                continue
            if normalized_path:
                attempted_paths.add(normalized_path)

            print(f"[BROWSER] Starting Chrome with {source_name}...")
            service = Service(
                executable_path=driver_path,
                log_output=log_path,
                service_args=["--verbose"],
            )
            driver = webdriver.Chrome(
                service=service,
                options=build_chrome_options(profile_dir),
            )
            driver._nclt_profile_dir = profile_dir
            driver._nclt_log_path = log_path
            return driver
        except Exception as exc:
            errors.append(f"{source_name}: {exc}")
            shutil.rmtree(profile_dir, ignore_errors=True)
            if not KEEP_BROWSER_LOGS:
                try:
                    os.remove(log_path)
                except OSError:
                    pass
            print(f"[WARN] Chrome startup failed with {source_name}; retrying...")
            time.sleep(2)

    details = "\n".join(f"  - {error}" for error in errors)
    raise RuntimeError(
        "Chrome could not start after all recovery attempts. "
        + (f"Startup logs are in {RUN_DOWNLOAD_DIR}.\n" if KEEP_BROWSER_LOGS else "Startup logs were removed after the failed attempts. Set NCLT_KEEP_BROWSER_LOGS=1 to retain them.\n")
        + details
    )


def close_driver(driver):
    profile_dir = getattr(driver, "_nclt_profile_dir", None)
    log_path = getattr(driver, "_nclt_log_path", None)
    try:
        driver.quit()
    finally:
        if profile_dir:
            time.sleep(0.5)
            shutil.rmtree(profile_dir, ignore_errors=True)
        if log_path and not KEEP_BROWSER_LOGS:
            try:
                os.remove(log_path)
            except OSError:
                pass


def clean_cell(text):
    if not text:
        return ""
    return re.sub(r"\s+", " ", repair_doubled_ocr(text)).strip()


def is_section_7(text):
    """Detect Section 7 without treating Section 59 Rule 7 as Section 7."""
    if not text:
        return False
    value = clean_cell(text).upper()
    explicit_patterns = [
        r"\b(?:SEC(?:TION)?|U/S|U\\S|S\.)\s*[-.:]?\s*7\b",
        r"\bIBC\s*(?:UNDER\s*)?(?:SEC(?:TION)?\s*)?7\b",
    ]
    if any(re.search(pattern, value) for pattern in explicit_patterns):
        return True
    if re.search(r"\bRULE\s*[-.:]?\s*7\b", value):
        return False
    implicit_patterns = [
        r"\b7\s*(?:OF\s+)?(?:THE\s+)?(?:IBC|INSOLVENCY|CODE)\b",
        r"\b7\s*IBC\b",
        r"(?=.*\b(?:IB|IBC|CP)\b).*/7/",
    ]
    return any(re.search(pattern, value) for pattern in implicit_patterns)


def auto_format_excel(file_path):
    print("  > Formatting Excel columns and wrapped text...")
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    banner_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    banner_font = Font(bold=True, color="002060", size=11)
    normal_font = Font(size=10)

    for row in worksheet.iter_rows():
        is_banner = any(
            cell.value and COURT_OPEN_MARKER in str(cell.value)
            for cell in row
        )
        for cell in row:
            cell.font = banner_font if is_banner else normal_font
            if is_banner:
                cell.fill = banner_fill
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column in worksheet.columns:
        worksheet.column_dimensions[column[0].column_letter].width = 35
    workbook.save(file_path)


def page_tables(page):
    tables = page.find_tables(LINE_TABLE_SETTINGS)
    if not tables or sum(len(table.rows) for table in tables) < 2:
        tables = page.find_tables(TEXT_TABLE_SETTINGS)
    return tables


def add_recovered_case_id(row, case_id):
    if not case_id or extract_case_id_from_cells(row):
        return False
    if len(row) > 1 and not row[1]:
        row[1] = case_id
    elif row and not row[0]:
        row[0] = case_id
    else:
        row.append(case_id)
    return True


def merge_case_rows(rows):
    """Merge vertically split subrows while preserving the table's columns."""
    if not rows:
        return []
    rows = sorted(
        enumerate(rows),
        key=lambda item: (
            0
            if "IN THE MATTER OF" in " ".join(item[1]).upper()
            else 1
            if re.search(r"\bMAIN\s+(?:CASE|MATTER|PETITION)\b", " ".join(item[1]), re.I)
            else 2,
            item[0],
        ),
    )
    rows = [row for _, row in rows]
    width = max(len(row) for row in rows)
    merged = []
    for column_index in range(width):
        values = []
        for row in rows:
            if column_index >= len(row):
                continue
            value = clean_cell(row[column_index])
            if value and value not in values:
                values.append(value)
        merged.append("\n".join(values))
    return merged


def table_case_blocks(table):
    """Yield complete case rows, including continuation subrows from merged cells."""
    extracted_rows = table.extract() or []
    current_rows = []
    current_bbox = None

    def flush():
        nonlocal current_rows, current_bbox
        if not current_rows:
            return None
        result = (merge_case_rows(current_rows), current_bbox)
        current_rows = []
        current_bbox = None
        return result

    for row_index, row in enumerate(extracted_rows):
        clean_row = [clean_cell(cell) for cell in (row or [])]
        if not any(clean_row):
            continue

        row_bbox = table.rows[row_index].bbox
        case_id = extract_case_id_from_cells(clean_row)
        row_text = " ".join(clean_row)
        section_7_row = is_section_7(row_text)
        is_indented_continuation = row_bbox[0] > table.bbox[0] + 2
        first_nonempty = next(
            (index for index, value in enumerate(clean_row) if value),
            len(clean_row),
        )
        is_blank_leading_continuation = first_nonempty >= 2

        if case_id:
            completed = flush()
            if completed:
                yield completed
            current_rows = [clean_row]
            current_bbox = row_bbox
        elif section_7_row:
            completed = flush()
            if completed:
                yield completed
            current_rows = [clean_row]
            current_bbox = row_bbox
        elif current_rows and (
            is_indented_continuation or is_blank_leading_continuation
        ):
            current_rows.append(clean_row)
        else:
            completed = flush()
            if completed:
                yield completed

    completed = flush()
    if completed:
        yield completed


def process_pdfs_to_excel(pdf_mapping, download_dir=None, run_timestamp=None):
    source_dir = download_dir or RUN_DOWNLOAD_DIR
    output_timestamp = run_timestamp or RUN_TIMESTAMP
    print("\n[STEP 5] EXTRACTING: Parsing all downloaded cause lists...")
    pdf_files = sorted(glob.glob(os.path.join(source_dir, "*.pdf")))
    if not pdf_files:
        raise RuntimeError(f"No PDF files found in {source_dir}.")

    all_master_rows = []
    parse_errors = []
    recovered_case_ids = 0

    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        court_name = pdf_mapping.get(
            filename,
            pdf_mapping.get(
                urllib.parse.unquote(filename),
                filename.replace(".pdf", ""),
            ),
        ).upper()
        print(f"  > Parsing: {court_name}")
        court_cases = []

        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    for table in page_tables(page):
                        for clean_row, row_bbox in table_case_blocks(table):
                            if not is_section_7(" ".join(clean_row)):
                                continue

                            if not extract_case_id_from_cells(clean_row):
                                recovered = recover_case_id_from_row(
                                    page,
                                    row_bbox,
                                    table.bbox,
                                )
                                if add_recovered_case_id(clean_row, recovered):
                                    recovered_case_ids += 1
                            court_cases.append(clean_row)
        except Exception as exc:
            parse_errors.append(f"{filename}: {type(exc).__name__}: {exc}")

        if court_cases:
            all_master_rows.append(
                [f"{COURT_OPEN_MARKER} {court_name} {COURT_CLOSE_MARKER}"]
            )
            all_master_rows.extend(court_cases)
            all_master_rows.extend([[], [], []])

    if parse_errors:
        raise RuntimeError("PDF parsing failed: " + " | ".join(parse_errors))
    if not all_master_rows:
        return None

    print(f"  > Recovered {recovered_case_ids} case IDs from merged PDF cells.")
    print("\n[STEP 6] SAVING: Generating the Master workbook...")
    output_filename = f"Consolidated_IBC7_Master_{output_timestamp}.xlsx"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    max_columns = max(len(row) for row in all_master_rows)
    normalized_rows = [
        row + [""] * (max_columns - len(row))
        for row in all_master_rows
    ]
    pd.DataFrame(normalized_rows).to_excel(
        output_path,
        index=False,
        header=False,
        engine="openpyxl",
    )
    auto_format_excel(output_path)
    return output_path


def find_next_results_url(driver):
    """Return the NCLT pager's next-page URL without matching footer controls."""
    from selenium.webdriver.common.by import By

    for selector in NEXT_PAGE_SELECTORS:
        for link in driver.find_elements(By.CSS_SELECTOR, selector):
            if link.get_attribute("aria-disabled") == "true":
                continue
            href = link.get_attribute("href")
            if href:
                return urllib.parse.urljoin(driver.current_url, href)
    return None


def chrome_download_name(filename, occurrence):
    """Predict Chrome's name when different PDF URLs share a filename."""
    if occurrence == 0:
        return filename
    stem, extension = os.path.splitext(filename)
    return f"{stem} ({occurrence}){extension}"


def parse_result_date(text):
    """Parse the NCLT result-table date, which is displayed as DD/MM/YYYY."""
    normalized = re.sub(r"\s+", " ", text or "").strip()
    match = re.search(
        r"(?<!\d)(\d{1,2})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{4})(?!\d)",
        normalized,
    )
    if not match:
        return None
    try:
        return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    except ValueError:
        return None


def prompt_for_target_date():
    """Ask the operator which NCLT result date should be downloaded."""
    print("[NCLT] Enter a date available on the NCLT cause-list page.")
    while True:
        entered_date = input("[NCLT] Enter the date to download (DD/MM/YYYY): ").strip()
        target_date = parse_result_date(entered_date)
        if target_date:
            print(
                f"[NCLT] Selected result date: "
                f"{target_date.strftime('%d/%m/%Y')}"
            )
            return target_date
        print("[NCLT] Invalid date. Please enter a valid date such as 03/08/2026.")


def extract_result_row_date(row):
    """Read the result date from the table's final date cell."""
    from selenium.webdriver.common.by import By

    columns = row.find_elements(By.TAG_NAME, "td")
    if not columns:
        return None

    # The live table is Title, Court, No. of Entries, PDF File, Date.
    candidates = [columns[-1].text]
    for element in row.find_elements(
        By.CSS_SELECTOR,
        "[class*='cause-date'], [class*='field-date'], [data-date]",
    ):
        candidates.insert(0, element.text or element.get_attribute("data-date"))

    for text in candidates:
        parsed = parse_result_date(text)
        if parsed:
            return parsed
    return None


def extract_result_court(row):
    """Extract the court label from current or older NCLT table layouts."""
    from selenium.webdriver.common.by import By

    columns = row.find_elements(By.TAG_NAME, "td")
    if not columns:
        return "UNKNOWN COURT"

    current_court = clean_cell(columns[1].text) if len(columns) > 1 else ""
    if current_court and re.search(
        r"\b(?:bench|court)\b", current_court, re.IGNORECASE
    ):
        return current_court

    legacy_parts = []
    for column in columns[2:4]:
        value = clean_cell(column.text)
        if value and re.search(r"\b(?:bench|court)\b", value, re.IGNORECASE):
            legacy_parts.append(value)
    if legacy_parts:
        return " - ".join(dict.fromkeys(legacy_parts))

    return current_court or "UNKNOWN COURT"


def find_latest_result_date(driver, wait):
    """Return the first valid date from the website's ordered result table."""
    from selenium.webdriver.common.by import By

    rows = wait.until(
        lambda current: current.find_elements(By.XPATH, "//table//tr[td]")
        or False
    )
    for row in rows:
        result_date = extract_result_row_date(row)
        if result_date:
            return result_date
    raise RuntimeError("The NCLT page loaded, but no result date was found.")


def download_all_result_pages(driver, target_date):
    """Find target_date in the ordered table and download its PDFs across pages."""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    wait = WebDriverWait(driver, 45)
    pdf_mapping = {}
    seen_pdf_urls = set()
    visited_page_urls = set()
    filename_counts = {}
    result_page = 1

    while True:
        wait.until(
            lambda current: current.execute_script(
                "return document.readyState"
            ) == "complete"
        )
        try:
            rows = wait.until(
                lambda current: current.find_elements(By.XPATH, "//table//tr[td]")
                or False
            )
        except Exception as exc:
            raise RuntimeError(
                f"No NCLT result rows loaded on result page {result_page}. "
                "Check that the NCLT result table loaded successfully."
            ) from exc

        current_page_url = driver.current_url
        if current_page_url in visited_page_urls:
            raise RuntimeError(
                f"NCLT pagination returned to an already visited page: {current_page_url}"
            )
        visited_page_urls.add(current_page_url)
        new_links = []
        stop_after_page = False
        target_date_found = False

        for row in rows:
            row_date = extract_result_row_date(row)
            if row_date is None:
                raise RuntimeError(
                    f"Could not read the date for an NCLT result row on page {result_page}."
                )
            if row_date != target_date and target_date_found:
                stop_after_page = True
                break
            if row_date != target_date:
                continue

            target_date_found = True

            row_links = row.find_elements(
                By.XPATH,
                ".//a[contains(translate(@href, 'PDF', 'pdf'), '.pdf')]",
            )
            for link in row_links:
                href = link.get_attribute("href")
                if not href or href in seen_pdf_urls:
                    continue

                court_name = extract_result_court(row)
                filename = urllib.parse.unquote(
                    os.path.basename(urllib.parse.urlsplit(href).path)
                )
                occurrence = filename_counts.get(filename, 0)
                expected_filename = chrome_download_name(filename, occurrence)
                filename_counts[filename] = occurrence + 1
                pdf_mapping[expected_filename] = court_name
                seen_pdf_urls.add(href)
                new_links.append(link)

        print(
            f"  > Result page {result_page}: {len(rows)} rows, "
            f"{len(new_links)} new for {target_date.strftime('%d/%m/%Y')}."
        )
        for link in new_links:
            driver.execute_script("arguments[0].click();", link)
            time.sleep(0.25)

        if stop_after_page:
            print("  > Reached a different result date; stopping pagination.")
            break

        next_page_url = find_next_results_url(driver)
        if not next_page_url:
            break
        if next_page_url in visited_page_urls:
            raise RuntimeError(
                f"NCLT next-page link loops to {next_page_url}."
            )

        result_page += 1
        print(f"  > Opening result page {result_page}...")
        driver.get(next_page_url)

    print(
        f"  > Collected {len(seen_pdf_urls)} unique PDFs from "
        f"{result_page} result page(s)."
    )
    if not seen_pdf_urls:
        raise RuntimeError(
            f"No NCLT PDF results were found for {target_date.strftime('%d/%m/%Y')}."
        )
    return pdf_mapping, len(seen_pdf_urls)


def wait_for_downloads(expected_pdfs):
    if expected_pdfs <= 0:
        raise RuntimeError("The NCLT result page contained no PDF links.")

    print(f"  > Waiting for {expected_pdfs} PDFs to finish downloading...")
    last_snapshot = None
    last_progress_at = time.time()
    while True:
        downloaded = glob.glob(os.path.join(RUN_DOWNLOAD_DIR, "*.pdf"))
        partials = glob.glob(os.path.join(RUN_DOWNLOAD_DIR, "*.crdownload"))
        if len(downloaded) >= expected_pdfs and not partials:
            print(f"  > All {expected_pdfs} PDFs downloaded successfully.")
            return

        snapshot = (
            len(downloaded),
            tuple(
                sorted(
                    (os.path.basename(path), os.path.getsize(path))
                    for path in partials
                    if os.path.exists(path)
                )
            ),
        )
        if snapshot != last_snapshot:
            last_snapshot = snapshot
            last_progress_at = time.time()
        elif time.time() - last_progress_at > 120:
            break
        time.sleep(2)

    downloaded = glob.glob(os.path.join(RUN_DOWNLOAD_DIR, "*.pdf"))
    raise RuntimeError(
        f"Download incomplete: received {len(downloaded)} of {expected_pdfs} PDFs."
    )


def main():
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    print(f"\n[INIT] RUN ID: {RUN_TIMESTAMP}")
    target_date = prompt_for_target_date()
    os.makedirs(RUN_DOWNLOAD_DIR, exist_ok=True)
    driver = setup_driver()
    pdf_mapping = {}

    try:
        driver.get(BASE_URL)
        wait = WebDriverWait(driver, 45)
        wait.until(
            EC.presence_of_element_located(
                (By.NAME, "field_nclt_benches_list_target_id")
            )
        )
        latest_date = find_latest_result_date(driver, wait)
        print(
            f"[NCLT] First date currently shown on the website: "
            f"{latest_date.strftime('%d/%m/%Y')}"
        )
        pdf_mapping, expected_pdfs = download_all_result_pages(driver, target_date)
        wait_for_downloads(expected_pdfs)
    finally:
        close_driver(driver)

    report_path = process_pdfs_to_excel(pdf_mapping)
    if not report_path:
        print("\n[OK] No Section 7 IBC matters were found for the selected website date.")
        return None

    print(f"\n[OK] Master Excel report generated: {os.path.basename(report_path)}")
    return report_path


if __name__ == "__main__":
    main()
